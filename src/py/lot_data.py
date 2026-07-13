"""Leakage-aware preparation utilities for old and new COTA workbooks.

Raw patient identifiers exist only in memory. Generated artifacts use
deterministic, source-scoped pseudonyms and never serialize a raw identifier.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
from collections import Counter, OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from textbook_algo_cota import SCT_TOKENS, normalize_drug


SCHEMA_VERSION = "2.0.0"
KEY_NAMESPACE = "lot-data-foundation-v2"
ADJUDICATED_SOURCE = "cota_adjudicated"
UNADJUDICATED_SOURCE = "cota_unadjudicated"


def find_project_root(start: str | Path | None = None) -> Path:
    """Find the repository root from an override, cwd, or this module."""
    override = os.environ.get("LOT_PROJECT_ROOT")
    candidates = [Path(override)] if override else []
    candidates.extend([Path(start or Path.cwd()), Path(__file__).parent])
    visited: set[Path] = set()
    for candidate in candidates:
        candidate = candidate.expanduser().resolve()
        for directory in (candidate, *candidate.parents):
            if directory in visited:
                continue
            visited.add(directory)
            if (directory / "src" / "py").is_dir() and (directory / "data").is_dir():
                return directory
    raise FileNotFoundError(
        "Could not find project root; pass --project-root or set LOT_PROJECT_ROOT"
    )


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def as_int(value: Any) -> int | None:
    if is_blank(value):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return int(number) if math.isfinite(number) and number.is_integer() else None


def iso_date(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def pseudonymous_key(source: str, raw_patient_id: Any) -> str:
    payload = f"{KEY_NAMESPACE}|patient|{source}|{str(raw_patient_id).strip()}"
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:20]
    prefix = "COTAOLD" if source == ADJUDICATED_SOURCE else "COTANEW"
    return f"{prefix}-{digest}"


def case_key(patient_key: str) -> str:
    digest = hashlib.sha256(
        f"{KEY_NAMESPACE}|blind|{patient_key}".encode("utf-8")
    ).hexdigest()[:20]
    return f"CASE-{digest}"


def parse_treatment_groups(
    text: Any, *, bracketed: bool = True
) -> tuple[list[frozenset[str]], list[str], int]:
    """Return ordered normalized regimen groups from one reconstructed string.

    The order of bracketed groups is preserved. SCT-only groups are removed and
    counted separately because the existing COTA algorithm treats SCT as a
    non-drug token.
    """
    if is_blank(text):
        return [], ["missing_treatment"], 0

    value = str(text).strip()
    flags: list[str] = []
    balanced = value.count("[") == value.count("]")
    if not balanced:
        flags.append("unbalanced_brackets")

    if bracketed and balanced and "[" in value:
        raw_groups = re.findall(r"\[([^\]]*)\]", value)
    else:
        if bracketed and "[" not in value:
            flags.append("unbracketed_treatment")
        raw_groups = [re.sub(r"[\[\]]", "", value)]

    groups: list[frozenset[str]] = []
    removed_non_treatment_groups = 0
    for raw_group in raw_groups:
        normalized = [normalize_drug(token) for token in raw_group.split(",")]
        tokens = [token for token in normalized if token]
        drugs = frozenset(token for token in tokens if token not in SCT_TOKENS)
        if drugs:
            groups.append(drugs)
        elif tokens and all(token in SCT_TOKENS for token in tokens):
            removed_non_treatment_groups += 1
        else:
            flags.append("empty_regimen_group")

    if not groups and not removed_non_treatment_groups:
        flags.append("empty_drug_set")
    return groups, sorted(set(flags)), removed_non_treatment_groups


def parse_treatment(text: Any, *, bracketed: bool) -> tuple[frozenset[str], list[str]]:
    """Compatibility helper returning the union of parsed regimen groups."""
    groups, flags, _ = parse_treatment_groups(text, bracketed=bracketed)
    return frozenset().union(*groups) if groups else frozenset(), flags


@dataclass
class TreatmentEvent:
    order: int
    drugs: frozenset[str]
    source_line_number: int | None = None
    source_group_index: int | None = None
    source_line_start_date: str | None = None
    source_line_end_date: str | None = None
    quality_flags: list[str] = field(default_factory=list)

    def evaluation_dict(self) -> dict[str, Any]:
        return {
            "order": self.order,
            "drugs": sorted(self.drugs),
            "source_line_number": self.source_line_number,
            "source_group_index": self.source_group_index,
            "source_line_start_date": self.source_line_start_date,
            "source_line_end_date": self.source_line_end_date,
            "quality_flags": sorted(set(self.quality_flags)),
        }

    def blind_dict(self) -> dict[str, Any]:
        return {"order": self.order, "drugs": sorted(self.drugs)}


@dataclass
class PatientRecord:
    source: str
    raw_patient_id: str
    trajectory: list[TreatmentEvent]
    diagnosis_date: str | None
    reviewer_a_lot: int | None
    reviewer_b_lot: int | None
    reviewer_consensus_lot: int | None
    reviewer_consensus_status: str
    cota_lot: int | None
    quality_metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def key(self) -> str:
        return pseudonymous_key(self.source, self.raw_patient_id)

    def evaluation_dict(self) -> dict[str, Any]:
        return {
            "schema_version": SCHEMA_VERSION,
            "patient_key": self.key,
            "source": self.source,
            "context": {"diagnosis_date": self.diagnosis_date},
            "trajectory": [event.evaluation_dict() for event in self.trajectory],
            "cota_lot": self.cota_lot,
            "reviewer_lot": {
                "reviewer_a": self.reviewer_a_lot,
                "reviewer_b": self.reviewer_b_lot,
                "consensus": self.reviewer_consensus_lot,
                "consensus_status": self.reviewer_consensus_status,
            },
            "quality_metadata": self.quality_metadata,
        }

    def blind_dict(self) -> dict[str, Any]:
        return {
            "schema_version": SCHEMA_VERSION,
            "case_key": case_key(self.key),
            "trajectory": [event.blind_dict() for event in self.trajectory],
            "context": {"diagnosis_date": self.diagnosis_date},
        }


@dataclass
class LoadStats:
    source: str
    sheet_names: list[str]
    selected_sheet: str
    observed_headers: list[str]
    missing_semantic_fields: list[str]
    raw_rows: int = 0
    patient_count: int = 0
    vendor_line_count: int = 0
    flattened_treatment_event_count: int = 0
    continuation_rows: int = 0
    same_patient_new_line_rows: int = 0
    orphan_continuation_rows: int = 0
    rows_without_patient_or_treatment: int = 0
    malformed_line_number_rows: int = 0

    def public_dict(self) -> dict[str, Any]:
        return dict(vars(self))


def consensus(a: int | None, b: int | None) -> tuple[int | None, str]:
    if a is None or b is None:
        return None, "missing_reviewer"
    if a != b:
        return None, "disagreement"
    return a, "agreed"


def _worksheet_with_headers(
    path: str | Path, sheet: str | None
) -> tuple[Any, list[str], list[str]]:
    # Non-read-only mode is intentional: some vendor files omit worksheet
    # dimensions, which makes openpyxl read-only iteration appear empty.
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet_names = list(workbook.sheetnames)
    selected = sheet or sheet_names[0]
    if selected not in workbook.sheetnames:
        raise ValueError(f"Sheet {selected!r} not found; available sheets: {sheet_names}")
    worksheet = workbook[selected]
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in first_row]
    return worksheet, sheet_names, headers


def _first(values: Iterable[Any]) -> Any:
    return next((value for value in values if not is_blank(value)), None)


def _load_cota_workbook(
    path: str | Path,
    *,
    source: str,
    sheet: str | None,
    adjudicated: bool,
) -> tuple[list[PatientRecord], LoadStats]:
    worksheet, sheet_names, headers = _worksheet_with_headers(path, sheet)
    columns = {name: index for index, name in enumerate(headers) if name}
    required = {
        "cpid", "line_of_therapy_c", "line_of_therapy_name",
        "date_start_line_of_therapy", "date_end_line_of_therapy",
    }
    missing = required - columns.keys()
    if missing:
        raise ValueError(f"COTA sheet missing required columns: {sorted(missing)}")
    reviewer_columns = {"Alpesh 1 LoT", "Alberto LOT"}
    if adjudicated and not reviewer_columns <= columns.keys():
        raise ValueError("Adjudicated COTA sheet is missing reviewer columns")

    missing_semantic = [] if adjudicated else [
        "reviewer_a_lot", "reviewer_b_lot", "reviewer_consensus_lot"
    ]
    stats = LoadStats(
        source=source,
        sheet_names=sheet_names,
        selected_sheet=worksheet.title,
        observed_headers=[header for header in headers if header],
        missing_semantic_fields=missing_semantic,
    )
    grouped: OrderedDict[tuple[str, int | None], dict[str, Any]] = OrderedDict()
    patients_in_order: OrderedDict[str, None] = OrderedDict()
    reviewer_values: dict[str, dict[str, list[int]]] = defaultdict(lambda: {"a": [], "b": []})
    diagnosis_dates: dict[str, list[Any]] = defaultdict(list)
    patient_quality: dict[str, Counter[str]] = defaultdict(Counter)
    current_patient: str | None = None
    current_line: int | None = None

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        stats.raw_rows += 1
        raw_id_value = row[columns["cpid"]]
        line_value = row[columns["line_of_therapy_c"]]
        treatment = row[columns["line_of_therapy_name"]]
        has_id = not is_blank(raw_id_value)
        has_line = not is_blank(line_value)
        has_treatment = not is_blank(treatment)

        if has_id:
            current_patient = str(raw_id_value).strip()
            patients_in_order.setdefault(current_patient, None)
            current_line = None
        elif current_patient is None:
            if has_treatment:
                stats.orphan_continuation_rows += 1
            else:
                stats.rows_without_patient_or_treatment += 1
            continue

        assert current_patient is not None
        if has_line:
            parsed_line = as_int(line_value)
            if parsed_line is None:
                stats.malformed_line_number_rows += 1
                patient_quality[current_patient]["malformed_line_number"] += 1
            current_line = parsed_line
            if not has_id:
                stats.same_patient_new_line_rows += 1
        elif not has_id and has_treatment:
            stats.continuation_rows += 1
        elif not has_id and not has_treatment:
            stats.rows_without_patient_or_treatment += 1
            continue

        if "diag_dt" in columns:
            diagnosis_dates[current_patient].append(row[columns["diag_dt"]])
        if adjudicated:
            a = as_int(row[columns["Alpesh 1 LoT"]])
            b = as_int(row[columns["Alberto LOT"]])
            if a is not None:
                reviewer_values[current_patient]["a"].append(a)
            if b is not None:
                reviewer_values[current_patient]["b"].append(b)

        key = (current_patient, current_line)
        group = grouped.setdefault(key, {
            "first_row": row_number,
            "parts": [],
            "continuations": 0,
            "start_dates": [],
            "end_dates": [],
        })
        if has_treatment:
            group["parts"].append(str(treatment))
        if not has_id and not has_line and has_treatment:
            group["continuations"] += 1
        group["start_dates"].append(row[columns["date_start_line_of_therapy"]])
        group["end_dates"].append(row[columns["date_end_line_of_therapy"]])

    by_patient: OrderedDict[str, list[tuple[int | None, dict[str, Any]]]] = OrderedDict(
        (raw_id, []) for raw_id in patients_in_order
    )
    for (raw_id, line_number), group in grouped.items():
        by_patient[raw_id].append((line_number, group))

    records: list[PatientRecord] = []
    for raw_id, source_lines in by_patient.items():
        source_lines.sort(key=lambda item: (
            item[0] is None,
            item[0] if item[0] is not None else item[1]["first_row"],
            item[1]["first_row"],
        ))
        trajectory: list[TreatmentEvent] = []
        parser_flags: Counter[str] = Counter()
        removed_non_treatment_groups = 0
        continuation_count = 0
        for line_number, source_line in source_lines:
            treatment = "".join(source_line["parts"])
            groups, flags, removed = parse_treatment_groups(treatment, bracketed=True)
            parser_flags.update(flags)
            removed_non_treatment_groups += removed
            continuation_count += source_line["continuations"]
            for group_index, drugs in enumerate(groups, start=1):
                trajectory.append(TreatmentEvent(
                    order=len(trajectory) + 1,
                    drugs=drugs,
                    source_line_number=line_number,
                    source_group_index=group_index,
                    source_line_start_date=iso_date(_first(source_line["start_dates"])),
                    source_line_end_date=iso_date(_first(source_line["end_dates"])),
                    quality_flags=flags,
                ))

        if adjudicated:
            values = reviewer_values[raw_id]
            reviewer_a = max(values["a"], default=None)
            reviewer_b = max(values["b"], default=None)
            reviewer_consensus, consensus_status = consensus(reviewer_a, reviewer_b)
        else:
            reviewer_a = reviewer_b = reviewer_consensus = None
            consensus_status = "unadjudicated"
        cota_values = [line for line, _ in source_lines if line is not None]
        quality_counts = patient_quality[raw_id]
        records.append(PatientRecord(
            source=source,
            raw_patient_id=raw_id,
            trajectory=trajectory,
            diagnosis_date=iso_date(_first(diagnosis_dates[raw_id])),
            reviewer_a_lot=reviewer_a,
            reviewer_b_lot=reviewer_b,
            reviewer_consensus_lot=reviewer_consensus,
            reviewer_consensus_status=consensus_status,
            cota_lot=max(cota_values, default=None),
            quality_metadata={
                "vendor_line_count": len(source_lines),
                "continuation_row_count": continuation_count,
                "parser_flag_counts": dict(sorted(parser_flags.items())),
                "removed_non_treatment_group_count": removed_non_treatment_groups,
                "source_quality_counts": dict(sorted(quality_counts.items())),
            },
        ))

    stats.patient_count = len(records)
    stats.vendor_line_count = sum(
        record.quality_metadata["vendor_line_count"] for record in records
    )
    stats.flattened_treatment_event_count = sum(len(record.trajectory) for record in records)
    worksheet.parent.close()
    return records, stats


def load_adjudicated_cota(
    path: str | Path, sheet: str = "Cota"
) -> tuple[list[PatientRecord], LoadStats]:
    return _load_cota_workbook(
        path, source=ADJUDICATED_SOURCE, sheet=sheet, adjudicated=True
    )


def load_new_cota(
    path: str | Path, sheet: str | None = None
) -> tuple[list[PatientRecord], LoadStats]:
    return _load_cota_workbook(
        path, source=UNADJUDICATED_SOURCE, sheet=sheet, adjudicated=False
    )


# Kept as an import-compatible name for callers of foundation v1.
load_cota = load_adjudicated_cota


def trajectory_signature(record: PatientRecord, *, collapse_consecutive: bool = False) -> str:
    sequence: list[tuple[str, ...]] = []
    for event in record.trajectory:
        drugs = tuple(sorted(event.drugs))
        if not drugs:
            continue
        if collapse_consecutive and sequence and drugs == sequence[-1]:
            continue
        sequence.append(drugs)
    return json.dumps(sequence, separators=(",", ":"), ensure_ascii=True)


def duplicate_summary(records: list[PatientRecord]) -> dict[str, int]:
    exact = Counter(trajectory_signature(record) for record in records)
    collapsed = Counter(
        trajectory_signature(record, collapse_consecutive=True) for record in records
    )
    consecutive_events = 0
    patients_with_consecutive = 0
    for record in records:
        count = sum(
            current.drugs == previous.drugs
            for previous, current in zip(record.trajectory, record.trajectory[1:])
        )
        consecutive_events += count
        patients_with_consecutive += count > 0
    return {
        "exact_duplicate_trajectory_groups": sum(count > 1 for count in exact.values()),
        "patients_in_exact_duplicate_trajectory_groups": sum(
            count for count in exact.values() if count > 1
        ),
        "collapsed_duplicate_trajectory_groups": sum(
            count > 1 for count in collapsed.values()
        ),
        "patients_in_collapsed_duplicate_trajectory_groups": sum(
            count for count in collapsed.values() if count > 1
        ),
        "patients_with_consecutive_duplicate_events": patients_with_consecutive,
        "consecutive_duplicate_event_count": consecutive_events,
    }


def malformed_summary(records: list[PatientRecord]) -> dict[str, Any]:
    flag_events: Counter[str] = Counter()
    patient_flags: Counter[str] = Counter()
    patients_with_flags = 0
    for record in records:
        counts = Counter(record.quality_metadata.get("parser_flag_counts", {}))
        flag_events.update(counts)
        if counts:
            patients_with_flags += 1
            patient_flags.update(counts.keys())
    return {
        "flag_occurrences": dict(sorted(flag_events.items())),
        "patient_counts_by_flag": dict(sorted(patient_flags.items())),
        "patients_with_malformed_records": patients_with_flags,
    }


class _UnionFind:
    def __init__(self, keys: Iterable[str]) -> None:
        self.parent = {key: key for key in keys}

    def find(self, key: str) -> str:
        parent = self.parent[key]
        if parent != key:
            self.parent[key] = self.find(parent)
        return self.parent[key]

    def union(self, left: str, right: str) -> None:
        left_root, right_root = self.find(left), self.find(right)
        if left_root != right_root:
            winner, loser = sorted((left_root, right_root))
            self.parent[loser] = winner


def build_exclusion_groups(records: list[PatientRecord]) -> list[dict[str, Any]]:
    """Connect records sharing exact/collapsed trajectories or cross-source IDs."""
    by_key = {record.key: record for record in records}
    union_find = _UnionFind(by_key)
    reasons: dict[tuple[str, str], set[str]] = defaultdict(set)

    def connect(grouped: dict[str, list[str]], reason: str) -> None:
        for members in grouped.values():
            if len(members) < 2:
                continue
            anchor = members[0]
            for member in members[1:]:
                union_find.union(anchor, member)
                reasons[tuple(sorted((anchor, member)))].add(reason)

    exact: dict[str, list[str]] = defaultdict(list)
    collapsed: dict[str, list[str]] = defaultdict(list)
    raw_ids: dict[str, list[str]] = defaultdict(list)
    for record in records:
        exact[trajectory_signature(record)].append(record.key)
        collapsed[trajectory_signature(record, collapse_consecutive=True)].append(record.key)
        raw_ids[record.raw_patient_id].append(record.key)
    connect(exact, "exact_trajectory")
    connect(collapsed, "collapsed_trajectory")
    # Raw IDs link records only when they come from more than one workbook.
    connect({
        raw_id: keys for raw_id, keys in raw_ids.items()
        if len({by_key[key].source for key in keys}) > 1
    }, "cross_workbook_patient_id")

    components: dict[str, list[str]] = defaultdict(list)
    for key in by_key:
        components[union_find.find(key)].append(key)

    output: list[dict[str, Any]] = []
    for members in components.values():
        members = sorted(members)
        digest = hashlib.sha256("|".join(members).encode("utf-8")).hexdigest()[:20]
        component_reasons: set[str] = set()
        for pair, pair_reasons in reasons.items():
            if pair[0] in members and pair[1] in members:
                component_reasons.update(pair_reasons)
        for key in members:
            record = by_key[key]
            output.append({
                "patient_key": key,
                "source": record.source,
                "exclusion_group": f"EXCL-{digest}",
                "group_reasons": sorted(component_reasons),
                "reviewer_consensus_lot": record.reviewer_consensus_lot,
            })
    return sorted(output, key=lambda row: row["patient_key"])


def assign_grouped_folds(
    records: list[PatientRecord], exclusion_rows: list[dict[str, Any]], n_folds: int = 5
) -> list[dict[str, Any]]:
    """Assign whole exclusion groups to folds without trajectory or ID leakage."""
    if n_folds < 2:
        raise ValueError("n_folds must be at least 2")
    by_key = {record.key: record for record in records}
    groups: dict[str, list[str]] = defaultdict(list)
    for row in exclusion_rows:
        groups[row["exclusion_group"]].append(row["patient_key"])

    fold_old_sizes = [0] * n_folds
    fold_all_sizes = [0] * n_folds
    fold_labels = [Counter() for _ in range(n_folds)]
    assignments: dict[str, int] = {}

    def group_sort(item: tuple[str, list[str]]) -> tuple[int, str]:
        group, keys = item
        old_count = sum(by_key[key].source == ADJUDICATED_SOURCE for key in keys)
        return (-old_count, group)

    for group, keys in sorted(groups.items(), key=group_sort):
        label_counts = Counter(
            by_key[key].reviewer_consensus_lot
            for key in keys
            if by_key[key].source == ADJUDICATED_SOURCE
        )
        old_count = sum(label_counts.values())
        if old_count:
            fold = min(
                range(n_folds),
                key=lambda candidate: (
                    sum(fold_labels[candidate][label] * count for label, count in label_counts.items()),
                    fold_old_sizes[candidate],
                    fold_all_sizes[candidate],
                    candidate,
                ),
            )
        else:
            fold = min(range(n_folds), key=lambda candidate: (fold_all_sizes[candidate], candidate))
        assignments[group] = fold
        fold_old_sizes[fold] += old_count
        fold_all_sizes[fold] += len(keys)
        fold_labels[fold].update(label_counts)

    return [
        {**row, "fold": assignments[row["exclusion_group"]]}
        for row in exclusion_rows
    ]


def overlap_public_summary(
    old: list[PatientRecord], new: list[PatientRecord], exclusion_rows: list[dict[str, Any]]
) -> dict[str, int]:
    old_ids = {record.raw_patient_id for record in old}
    new_ids = {record.raw_patient_id for record in new}
    old_exact = Counter(trajectory_signature(record) for record in old)
    new_exact = Counter(trajectory_signature(record) for record in new)
    old_collapsed = Counter(
        trajectory_signature(record, collapse_consecutive=True) for record in old
    )
    new_collapsed = Counter(
        trajectory_signature(record, collapse_consecutive=True) for record in new
    )
    groups = Counter(row["exclusion_group"] for row in exclusion_rows)
    multi_record_groups = {group for group, count in groups.items() if count > 1}
    cross_sources: dict[str, set[str]] = defaultdict(set)
    for row in exclusion_rows:
        cross_sources[row["exclusion_group"]].add(row["source"])
    return {
        "cross_workbook_raw_id_overlap_count": len(old_ids & new_ids),
        "shared_exact_trajectory_signature_count": len(old_exact.keys() & new_exact.keys()),
        "shared_exact_trajectory_patient_pair_count": sum(
            old_exact[sig] * new_exact[sig] for sig in old_exact.keys() & new_exact.keys()
        ),
        "shared_collapsed_trajectory_signature_count": len(
            old_collapsed.keys() & new_collapsed.keys()
        ),
        "shared_collapsed_trajectory_patient_pair_count": sum(
            old_collapsed[sig] * new_collapsed[sig]
            for sig in old_collapsed.keys() & new_collapsed.keys()
        ),
        "exclusion_group_count": len(groups),
        "multi_record_exclusion_group_count": len(multi_record_groups),
        "records_in_multi_record_exclusion_groups": sum(
            groups[group] for group in multi_record_groups
        ),
        "cross_workbook_exclusion_group_count": sum(
            len(sources) > 1 for sources in cross_sources.values()
        ),
    }


def schema_difference_summary(old_stats: LoadStats, new_stats: LoadStats) -> dict[str, Any]:
    old_headers = set(old_stats.observed_headers)
    new_headers = set(new_stats.observed_headers)
    return {
        "shared_columns": sorted(old_headers & new_headers),
        "adjudicated_only_columns": sorted(old_headers - new_headers),
        "unadjudicated_only_columns": sorted(new_headers - old_headers),
        "unadjudicated_missing_semantic_fields": new_stats.missing_semantic_fields,
    }


def reconstruct_vendor_line_sequence(record: dict[str, Any]) -> list[frozenset[str]]:
    """Recombine private flattened groups for the unchanged baseline algorithm."""
    grouped: OrderedDict[int | None, set[str]] = OrderedDict()
    for event in record["trajectory"]:
        line_number = event["source_line_number"]
        grouped.setdefault(line_number, set()).update(event["drugs"])
    return [frozenset(drugs) for drugs in grouped.values()]


def write_evaluation_jsonl(path: str | Path, records: list[PatientRecord]) -> None:
    _write_jsonl_dicts(path, [record.evaluation_dict() for record in records], "patient_key")


def write_blind_jsonl(path: str | Path, records: list[PatientRecord]) -> None:
    _write_jsonl_dicts(path, [record.blind_dict() for record in records], "case_key")


def _write_jsonl_dicts(path: str | Path, rows: list[dict[str, Any]], key: str) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in sorted(rows, key=lambda item: item[key]):
            handle.write(json.dumps(row, sort_keys=True, separators=(",", ":")))
            handle.write("\n")


def write_jsonl_rows(path: str | Path, rows: list[dict[str, Any]], key: str) -> None:
    _write_jsonl_dicts(path, rows, key)


def write_json(path: str | Path, value: Any) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(value, handle, indent=2, sort_keys=True)
        handle.write("\n")


def assert_aggregate_only_report(value: dict[str, Any]) -> None:
    """Reject patient-level identifiers/case collections in a public report."""
    if value.get("report_scope") != "aggregate_only":
        raise ValueError("Public report must declare report_scope=aggregate_only")
    forbidden_keys = {
        "patient_key", "case_key", "raw_patient_id", "cpid_value", "patid_value",
        "both_agree_but_wrong_cases", "patient_cases", "records",
    }

    def walk(item: Any) -> None:
        if isinstance(item, dict):
            overlap = forbidden_keys & item.keys()
            if overlap:
                raise ValueError(f"Patient-level fields forbidden in public report: {sorted(overlap)}")
            for nested in item.values():
                walk(nested)
        elif isinstance(item, list):
            for nested in item:
                walk(nested)

    walk(value)


def write_fold_manifest(path: str | Path, rows: list[dict[str, Any]]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["patient_key", "fold", "exclusion_group", "reviewer_consensus_lot"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(sorted(({
            name: row[name] for name in fieldnames
        } for row in rows if row["source"] == ADJUDICATED_SOURCE), key=lambda row: row["patient_key"]))


def read_jsonl(path: str | Path) -> list[dict[str, Any]]:
    with Path(path).open(encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


read_public_jsonl = read_jsonl
