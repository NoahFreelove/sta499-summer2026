
"""
Color-code COTA misclassified rows by misclassification pattern.

FIXED VERSION:
  - Does NOT call rows "same_exact_regimen_repeated" just because the UNIQUE drug set is the same.
  - Compares regimens at three levels:
      1) normalized full regimen string
      2) bracket-level regimen chunks / phases
      3) unique drug set
  - Adds a separate pattern for cases like:
      [carfilzomib, daratumumab, dexamethasone], [daratumumab]
      vs
      [carfilzomib, daratumumab, dexamethasone]
    These have the same unique drug set but different regimen structure.
"""

from __future__ import annotations

from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
INPUT_PATH = BASE_DIR / "Output" / "COTA_misclassified_rows_UPD.xlsx"
OUTPUT_PATH = BASE_DIR / "Output" / "COTA_misclassified_patterns_colored.xlsx"
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# -----------------------------------------------------------------------------
# Colors / flags
# -----------------------------------------------------------------------------
PATTERN_COLORS = {
    "same_exact_regimen_repeated": "D9EAD3",
    "same_drug_set_different_regimen_structure": "B6D7A8",
    "regimen_phase_drop_or_deescalation": "CFE2F3",
    "drug_drop_or_deescalation": "FFF2CC",
    "family_drop_or_deescalation": "FCE5CD",
    "re_expansion_to_prior_regimen": "D9EAF7",
    "new_drug_addition_requires_review": "F4CCCC",
    "same_family_drug_substitution": "EADCF8",
    "complex_transplant_or_cell_therapy": "D9D2E9",
    "complex_multi_drug_change_requires_review": "E6B8AF",
    "unknown_pattern_requires_review": "D9D9D9",
}

FIXABLE_PATTERNS = {
    "same_exact_regimen_repeated",
    "same_drug_set_different_regimen_structure",
    "regimen_phase_drop_or_deescalation",
    "drug_drop_or_deescalation",
    "family_drop_or_deescalation",
    "re_expansion_to_prior_regimen",
}

COMPLEX_THERAPY_KEYWORDS = {
    "transplant", "car-t", "cart", "autologous sct", "stem_cell_boost",
    "melphalan", "busulfan", "conditioning"
}

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_category(value) -> str:
    text = clean_text(value)
    return text.replace(" (Not in provided Fiona's category list)", "").strip()


def clean_token(value: str) -> str:
    token = clean_text(value).lower()
    token = token.strip().strip("[]'").strip('"')
    token = token.replace("[", "").replace("]", "")
    token = re.sub(r"\s+", " ", token).strip()
    return token


def split_list_string(value) -> list[str]:
    """Split values like 'a, b, c' or 'A + B + C' into clean lowercase tokens."""
    text = clean_text(value)
    if not text:
        return []
    text = text.replace(" (Not in provided Fiona's category list)", "")
    raw_parts = text.split(" + ") if " + " in text else text.split(",")
    parts = []
    for part in raw_parts:
        token = clean_token(part)
        if token and token not in parts:
            parts.append(token)
    return parts


def normalize_regimen_string(value) -> str:
    """
    Normalize the full regimen string while preserving bracket/chunk structure.
    This is used for TRUE exact repeat detection.
    """
    text = clean_text(value).lower()
    if not text:
        return ""
    text = text.replace("[ ", "[").replace(" ]", "]")
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\]\s*,\s*\[", "], [", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_regimen_chunks(value) -> list[tuple[str, ...]]:
    """
    Parse bracket-level regimen chunks.

    Example:
      '[carfilzomib, daratumumab, dexamethasone], [daratumumab]'
    returns:
      [('carfilzomib','daratumumab','dexamethasone'), ('daratumumab',)]

    We keep chunks because:
      A+B+C, B is NOT structurally the same as A+B+C.
    """
    text = normalize_regimen_string(value)
    if not text:
        return []

    bracket_chunks = re.findall(r"\[([^\]]+)\]", text)
    if not bracket_chunks:
        # Fallback for imperfect strings without balanced brackets.
        bracket_chunks = [text.replace("[", "").replace("]", "")]

    chunks: list[tuple[str, ...]] = []
    for chunk in bracket_chunks:
        drugs = []
        for raw in chunk.split(","):
            drug = clean_token(raw)
            if drug and drug not in drugs:
                drugs.append(drug)
        if drugs:
            chunks.append(tuple(drugs))
    return chunks


def chunk_list_to_str(chunks: list[tuple[str, ...]]) -> str:
    return " | ".join("[" + ", ".join(chunk) + "]" for chunk in chunks)


def get_drug_set_from_chunks(chunks: list[tuple[str, ...]]) -> set[str]:
    return {drug for chunk in chunks for drug in chunk}


def get_drug_set(row: pd.Series) -> set[str]:
    # Prefer chunk parsing from line_of_therapy_name because parsed_drugs loses structure.
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    if chunks:
        return get_drug_set_from_chunks(chunks)
    if "parsed_drugs" in row and clean_text(row.get("parsed_drugs")):
        return set(split_list_string(row.get("parsed_drugs")))
    return set()


def get_regimen_signature(row: pd.Series) -> dict:
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    return {
        "normalized_regimen": normalize_regimen_string(row.get("line_of_therapy_name")),
        "chunks": chunks,
        "chunk_sets": [frozenset(chunk) for chunk in chunks],
        "drug_set": get_drug_set_from_chunks(chunks),
    }


def get_family_set(row: pd.Series) -> set[str]:
    if "mapped_family_classes" in row and clean_text(row.get("mapped_family_classes")):
        return set(split_list_string(row.get("mapped_family_classes")))
    return set(split_list_string(clean_category(row.get("reconstructed_family_combination"))))


def has_complex_therapy(drug_set: set[str], family_set: set[str]) -> bool:
    combined = " ".join(sorted(drug_set | family_set)).lower()
    return any(keyword in combined for keyword in COMPLEX_THERAPY_KEYWORDS)


def set_to_str(values: set[str]) -> str:
    return ", ".join(sorted(values))


def chunk_sets_subset(current_chunk_sets: list[frozenset], previous_chunk_sets: list[frozenset]) -> bool:
    """True if every current regimen chunk appears inside the previous regimen chunks."""
    if not current_chunk_sets or not previous_chunk_sets:
        return False
    return all(chunk in previous_chunk_sets for chunk in current_chunk_sets)


def any_current_chunk_subset_of_previous(current_chunk_sets: list[frozenset], previous_chunk_sets: list[frozenset]) -> bool:
    """
    Detect within-regimen de-escalation, e.g.
      previous: [A,B,C], [B,C]
      current:  [A,B,C], [B]
    Unique drug set may be identical, but one phase/chunk shrank.
    """
    if not current_chunk_sets or not previous_chunk_sets:
        return False
    for cur in current_chunk_sets:
        if cur in previous_chunk_sets:
            continue
        if any(cur.issubset(prev) for prev in previous_chunk_sets):
            return True
    return False


def classify_row(row: pd.Series, prev_row: pd.Series | None, earlier_same_doctor_rows: pd.DataFrame) -> dict:
    cur_sig = get_regimen_signature(row)
    current_drugs = cur_sig["drug_set"]
    current_chunks = cur_sig["chunks"]
    current_chunk_sets = cur_sig["chunk_sets"]
    current_families = get_family_set(row)

    prev_sig = {"normalized_regimen": "", "chunks": [], "chunk_sets": [], "drug_set": set()}
    previous_drugs = set()
    previous_chunks = []
    previous_chunk_sets = []
    previous_families = set()

    if prev_row is not None:
        prev_sig = get_regimen_signature(prev_row)
        previous_drugs = prev_sig["drug_set"]
        previous_chunks = prev_sig["chunks"]
        previous_chunk_sets = prev_sig["chunk_sets"]
        previous_families = get_family_set(prev_row)

    added_drugs = current_drugs - previous_drugs
    removed_drugs = previous_drugs - current_drugs
    added_families = current_families - previous_families
    removed_families = previous_families - current_families

    # Look backward inside the same doctor LoT for true re-expansion/cycling.
    seen_same_full_regimen_earlier = False
    seen_same_chunk_structure_earlier = False
    seen_same_drug_set_earlier = False
    seen_same_family_set_earlier = False

    if not earlier_same_doctor_rows.empty:
        for _, earlier in earlier_same_doctor_rows.iterrows():
            earlier_sig = get_regimen_signature(earlier)
            if earlier_sig["normalized_regimen"] == cur_sig["normalized_regimen"] and cur_sig["normalized_regimen"]:
                seen_same_full_regimen_earlier = True
            if earlier_sig["chunks"] == current_chunks and current_chunks:
                seen_same_chunk_structure_earlier = True
            if earlier_sig["drug_set"] == current_drugs and current_drugs:
                seen_same_drug_set_earlier = True
            if get_family_set(earlier) == current_families and current_families:
                seen_same_family_set_earlier = True

    complex_case = has_complex_therapy(current_drugs, current_families) or has_complex_therapy(previous_drugs, previous_families)

    if not previous_drugs and not previous_families:
        pattern = "unknown_pattern_requires_review"
        explanation = "Previous patient row was not found, so the change cannot be compared safely."

    elif cur_sig["normalized_regimen"] == prev_sig["normalized_regimen"] and cur_sig["normalized_regimen"]:
        pattern = "same_exact_regimen_repeated"
        explanation = "COTA started a new LoT even though the full normalized regimen string is identical to the previous patient row."

    elif current_chunks == previous_chunks and current_chunks:
        pattern = "same_exact_regimen_repeated"
        explanation = "COTA started a new LoT even though the bracket-level regimen structure is identical to the previous patient row."

    elif current_drugs == previous_drugs and current_chunks != previous_chunks:
        if chunk_sets_subset(current_chunk_sets, previous_chunk_sets):
            pattern = "regimen_phase_drop_or_deescalation"
            explanation = "Unique drug set is the same, but one regimen phase/chunk was removed. This is not an exact repeat; it looks like de-escalation within the same doctor LoT."
        elif any_current_chunk_subset_of_previous(current_chunk_sets, previous_chunk_sets):
            pattern = "same_drug_set_different_regimen_structure"
            explanation = "Unique drug set is the same, but bracket-level regimen structure changed; one phase appears smaller/different. Not an exact repeat."
        else:
            pattern = "same_drug_set_different_regimen_structure"
            explanation = "Unique drug set is identical, but the full regimen/chunk structure changed. This should not be labeled same_exact_regimen_repeated."

    elif (seen_same_full_regimen_earlier or seen_same_chunk_structure_earlier) and (added_drugs or removed_drugs):
        pattern = "re_expansion_to_prior_regimen"
        explanation = "The exact regimen/chunk structure appeared earlier within the same doctor LoT; this looks like shrink/re-expand cycling rather than a true new LoT."

    elif current_drugs and previous_drugs and current_drugs.issubset(previous_drugs):
        pattern = "drug_drop_or_deescalation"
        explanation = "Current unique drug set is a subset of the previous unique drug set; COTA appears to have over-split after a drug drop/de-escalation."

    elif current_families and previous_families and current_families.issubset(previous_families):
        pattern = "family_drop_or_deescalation"
        explanation = "Current family combination is a subset of the previous family combination; likely de-escalation rather than a true new LoT."

    elif complex_case:
        pattern = "complex_transplant_or_cell_therapy"
        explanation = "Transition involves transplant/CAR-T/conditioning-related therapy; keep for human review."

    elif current_families == previous_families and current_drugs != previous_drugs:
        pattern = "same_family_drug_substitution"
        explanation = "Drug names changed while the family combination stayed the same; this may be a within-family substitution and needs review."

    elif added_drugs and not removed_drugs:
        if seen_same_full_regimen_earlier or seen_same_chunk_structure_earlier or seen_same_drug_set_earlier or seen_same_family_set_earlier:
            pattern = "re_expansion_to_prior_regimen"
            explanation = "Regimen/drug/family combination reappears within the same doctor LoT; likely add-back/re-expansion."
        else:
            pattern = "new_drug_addition_requires_review"
            explanation = "COTA over-split when a drug was added, but new additions can be clinically meaningful, so this needs human review."

    elif added_drugs and removed_drugs:
        pattern = "complex_multi_drug_change_requires_review"
        explanation = "Both added and removed drugs are present; this is a complex swap/change and should be reviewed manually."

    else:
        pattern = "unknown_pattern_requires_review"
        explanation = "No simple deterministic pattern was detected."

    review_flag = 1 if pattern in FIXABLE_PATTERNS else 2
    suggested_corrected_lot = row.get("doctor_lot_numeric_for_transition") if review_flag == 1 else ""

    return {
        "misclassification_pattern": pattern,
        "pattern_explanation": explanation,
        "review_flag": review_flag,
        "suggested_corrected_lot": suggested_corrected_lot,
        "previous_line_of_therapy_name": clean_text(prev_row.get("line_of_therapy_name")) if prev_row is not None else "",
        "previous_reconstructed_family_combination": clean_category(prev_row.get("reconstructed_family_combination")) if prev_row is not None else "",
        "current_regimen_normalized": cur_sig["normalized_regimen"],
        "previous_regimen_normalized": prev_sig["normalized_regimen"],
        "current_regimen_chunks": chunk_list_to_str(current_chunks),
        "previous_regimen_chunks": chunk_list_to_str(previous_chunks),
        "current_drugs_clean": set_to_str(current_drugs),
        "previous_drugs_clean": set_to_str(previous_drugs),
        "added_drugs_vs_previous": set_to_str(added_drugs),
        "removed_drugs_vs_previous": set_to_str(removed_drugs),
        "current_families_clean": set_to_str(current_families),
        "previous_families_clean": set_to_str(previous_families),
        "added_families_vs_previous": set_to_str(added_families),
        "removed_families_vs_previous": set_to_str(removed_families),
    }

# -----------------------------------------------------------------------------
# Load data
# -----------------------------------------------------------------------------
xl = pd.ExcelFile(INPUT_PATH)
if "Misclassified_Rows" not in xl.sheet_names:
    raise ValueError("Input workbook must contain a 'Misclassified_Rows' sheet.")

misclassified = pd.read_excel(INPUT_PATH, sheet_name="Misclassified_Rows")
misclassified.columns = misclassified.columns.str.strip()

if "All_COTA_With_Transitions" in xl.sheet_names:
    all_rows = pd.read_excel(INPUT_PATH, sheet_name="All_COTA_With_Transitions")
    all_rows.columns = all_rows.columns.str.strip()
else:
    all_rows = misclassified.copy()

required = ["cpid", "_original_row_order", "doctor_lot_numeric_for_transition"]
missing = [col for col in required if col not in all_rows.columns]
if missing:
    raise ValueError(f"Missing required columns in input workbook: {missing}")

all_rows = all_rows.sort_values(["cpid", "_original_row_order"]).copy()
misclassified = misclassified.sort_values(["cpid", "_original_row_order"]).copy()

all_rows["_row_key"] = all_rows["cpid"].astype(str) + "__" + all_rows["_original_row_order"].astype(str)
misclassified["_row_key"] = misclassified["cpid"].astype(str) + "__" + misclassified["_original_row_order"].astype(str)

# -----------------------------------------------------------------------------
# Classify each misclassified row
# -----------------------------------------------------------------------------
classified_records = []

for _, row in misclassified.iterrows():
    patient_rows = all_rows[all_rows["cpid"].astype(str).eq(str(row["cpid"]))].copy()
    patient_rows = patient_rows.sort_values("_original_row_order")

    earlier_rows = patient_rows[patient_rows["_original_row_order"] < row["_original_row_order"]]
    prev_row = earlier_rows.iloc[-1] if not earlier_rows.empty else None

    same_doctor_lot = row.get("doctor_lot_numeric_for_transition")
    earlier_same_doctor_rows = earlier_rows[
        earlier_rows["doctor_lot_numeric_for_transition"].eq(same_doctor_lot)
    ]

    classified_records.append(classify_row(row, prev_row, earlier_same_doctor_rows))

classification_df = pd.DataFrame(classified_records)
result = pd.concat([misclassified.reset_index(drop=True), classification_df], axis=1)

preferred_columns = [
    "misclassification_pattern",
    "review_flag",
    "suggested_corrected_lot",
    "pattern_explanation",
    "cpid",
    "line_of_therapy_c",
    "doctor_lot_numeric_for_transition",
    "previous_cota_lot_numeric",
    "previous_doctor_lot_numeric",
    "cota_minus_doctor_lot_shift",
    "reconstructed_family_combination",
    "line_of_therapy_name",
    "previous_line_of_therapy_name",
    "current_regimen_chunks",
    "previous_regimen_chunks",
    "current_regimen_normalized",
    "previous_regimen_normalized",
    "parsed_drugs",
    "current_drugs_clean",
    "previous_drugs_clean",
    "added_drugs_vs_previous",
    "removed_drugs_vs_previous",
    "current_families_clean",
    "previous_families_clean",
    "added_families_vs_previous",
    "removed_families_vs_previous",
    "transition_alignment_status",
    "discontinue_reason",
    "date_start_line_of_therapy",
    "date_end_line_of_therapy",
]
preferred_columns = [col for col in preferred_columns if col in result.columns]
remaining_columns = [col for col in result.columns if col not in preferred_columns]
result = result[preferred_columns + remaining_columns]

pattern_summary = (
    result.groupby(["misclassification_pattern", "review_flag"], dropna=False)
    .size()
    .reset_index(name="row_count")
    .sort_values("row_count", ascending=False)
)

category_pattern_summary = (
    result.groupby(["reconstructed_family_combination", "misclassification_pattern", "review_flag"], dropna=False)
    .size()
    .reset_index(name="row_count")
    .sort_values(["row_count", "reconstructed_family_combination"], ascending=[False, True])
)

fixable = result[result["review_flag"].eq(1)].copy()
human_review = result[result["review_flag"].eq(2)].copy()

legend_meaning = {
    "same_exact_regimen_repeated": "Full normalized regimen string/chunk structure is identical to previous row; likely safe auto-merge.",
    "same_drug_set_different_regimen_structure": "Same unique drugs, but bracket-level regimen structure differs; not an exact repeat.",
    "regimen_phase_drop_or_deescalation": "A regimen phase/chunk was removed or reduced while unique drugs may still look the same.",
    "drug_drop_or_deescalation": "Current unique drug set is subset of previous unique drug set; likely safe auto-merge.",
    "family_drop_or_deescalation": "Current family set is subset of previous family set; possible auto-merge.",
    "re_expansion_to_prior_regimen": "Exact regimen/chunk/drug/family combination appeared earlier in same doctor LoT; likely shrink/re-expand over-split.",
    "new_drug_addition_requires_review": "Drug was added; could be clinically meaningful, so review.",
    "same_family_drug_substitution": "Family stayed same but drug changed; review.",
    "complex_transplant_or_cell_therapy": "Transplant/CAR-T/conditioning-related case; review.",
    "complex_multi_drug_change_requires_review": "Both additions and removals; review.",
    "unknown_pattern_requires_review": "No safe rule detected; review.",
}

legend = pd.DataFrame([
    {
        "misclassification_pattern": pattern,
        "review_flag": 1 if pattern in FIXABLE_PATTERNS else 2,
        "color_hex": "#" + color,
        "meaning": legend_meaning.get(pattern, ""),
    }
    for pattern, color in PATTERN_COLORS.items()
])

# -----------------------------------------------------------------------------
# Write workbook
# -----------------------------------------------------------------------------
with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    result.to_excel(writer, sheet_name="Misclassified_Patterns", index=False)
    pattern_summary.to_excel(writer, sheet_name="Pattern_Summary", index=False)
    category_pattern_summary.to_excel(writer, sheet_name="Category_Pattern_Summary", index=False)
    fixable.to_excel(writer, sheet_name="Flag_1_Fixable", index=False)
    human_review.to_excel(writer, sheet_name="Flag_2_Human_Review", index=False)
    legend.to_excel(writer, sheet_name="Legend", index=False)

# -----------------------------------------------------------------------------
# Apply color formatting
# -----------------------------------------------------------------------------
wb = load_workbook(OUTPUT_PATH)
thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        header = str(ws.cell(row=1, column=col_idx).value or "")
        max_len = len(header)
        for cell in column_cells[1:80]:
            max_len = max(max_len, len(str(cell.value or "")))
        width = min(max(max_len + 2, 10), 45)
        if header in {
            "pattern_explanation", "line_of_therapy_name", "previous_line_of_therapy_name",
            "reconstructed_family_combination", "current_regimen_chunks", "previous_regimen_chunks"
        }:
            width = 42
        ws.column_dimensions[get_column_letter(col_idx)].width = width

ws = wb["Misclassified_Patterns"]
headers = [cell.value for cell in ws[1]]
pattern_col_idx = headers.index("misclassification_pattern") + 1

for row_idx in range(2, ws.max_row + 1):
    pattern = ws.cell(row=row_idx, column=pattern_col_idx).value
    fill_color = PATTERN_COLORS.get(pattern, PATTERN_COLORS["unknown_pattern_requires_review"])
    fill = PatternFill("solid", fgColor=fill_color)
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

ws_legend = wb["Legend"]
legend_headers = [cell.value for cell in ws_legend[1]]
legend_pattern_col = legend_headers.index("misclassification_pattern") + 1
for row_idx in range(2, ws_legend.max_row + 1):
    pattern = ws_legend.cell(row=row_idx, column=legend_pattern_col).value
    fill_color = PATTERN_COLORS.get(pattern, PATTERN_COLORS["unknown_pattern_requires_review"])
    fill = PatternFill("solid", fgColor=fill_color)
    for col_idx in range(1, ws_legend.max_column + 1):
        ws_legend.cell(row=row_idx, column=col_idx).fill = fill

wb.save(OUTPUT_PATH)

print(f"Saved: {OUTPUT_PATH}")
print(f"Total misclassified rows classified: {len(result)}")
print(f"Flag 1 likely script-fixable rows: {len(fixable)}")
print(f"Flag 2 human-review rows: {len(human_review)}")
print("\nPattern counts:")
print(pattern_summary.to_string(index=False))
