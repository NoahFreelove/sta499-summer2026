"""
Script to test how the COTA algorithm works.

TO-DO:
- implement the lot counting rules
- change the output to be in a copy of the input file
"""

from __future__ import annotations

from pathlib import Path
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
# Written by preprocessing_tester.py. The legacy COTA_preprocessing_newrules.py
# workbook is still accepted as a fallback (see sheet resolution below).
INPUT_PATH = BASE_DIR / "Output" / "COTA_cleaned.xlsx"
LEGACY_INPUT_PATH = BASE_DIR / "Output" / "COTA_misclassified_rows_UPD_newrules.xlsx"
OUTPUT_PATH = BASE_DIR / "Output" / "COTA_misclassified_patterns_colored_newrules.xlsx"
LOT_COUNTS_PATH = BASE_DIR / "Output" / "lot_counts_with_rules.xlsx"
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# Statuses that count as a misclassification, matching the filter that
# COTA_preprocessing_newrules.py used to build its "Misclassified_Rows" sheet.
MISCLASSIFIED_STATUSES = (
    "candidate_misclassification_cota_over_split",
    "candidate_misclassification_cota_under_split",
)

# -----------------------------------------------------------------------------
# Colors / flags
# -----------------------------------------------------------------------------
PATTERN_COLORS = {
    "same_exact_regimen_repeated": "D9EAD3",  # Green - safe merge
    "same_drug_set_different_regimen_structure": "B6D7A8",  # Light green
    "regimen_phase_drop_or_deescalation": "CFE2F3",  # Light blue
    "drug_drop_or_deescalation": "FFF2CC",  # Light yellow
    "family_drop_or_deescalation": "FCE5CD",  # Light orange
    "re_expansion_to_prior_regimen": "D9EAF7",  # Light steel blue
    "new_drug_addition_requires_review": "F4CCCC",  # Light red
    "same_family_drug_substitution": "EADCF8",  # Light purple
    "complex_transplant_or_cell_therapy": "D9D2E9",  # Medium purple
    "complex_multi_drug_change_requires_review": "E6B8AF",  # Salmon
    "unknown_pattern_requires_review": "D9D9D9",  # Grey
    # Rule engine specific patterns
    "p1_confirmed_progression": "FF9999",  # Dark red
    "p1b_planned_sequential": "99CCFF",  # Bright blue
    "rule1_steroid_only_absorbed": "CCFFCC",  # Light green
    "mandatory_drug_planned_triplet": "99FF99",  # Medium green
    "rules_2_3_identical_active_drugs": "CCCCFF",  # Light periwinkle
    "pre_asct_reinduction": "FFCC99",  # Light orange
    "asct_rule_post_transplant": "FFCCCC",  # Light red
    "car_t_rule": "FF99CC",  # Pink
    "p3_maintenance_after_combination": "CCFF99",  # Yellow-green
    "default_merge": "E6E6E6",  # Light grey
}

FIXABLE_PATTERNS = {
    "same_exact_regimen_repeated",
    "same_drug_set_different_regimen_structure",
    "regimen_phase_drop_or_deescalation",
    "drug_drop_or_deescalation",
    "family_drop_or_deescalation",
    "re_expansion_to_prior_regimen",
    "p1b_planned_sequential",
    "rule1_steroid_only_absorbed",
    "mandatory_drug_planned_triplet",
    "rules_2_3_identical_active_drugs",
    "pre_asct_reinduction",
    "asct_rule_post_transplant",
    "p3_maintenance_after_combination",
    "default_merge",
}

COMPLEX_THERAPY_KEYWORDS = {
    "transplant", "car-t", "cart", "autologous sct", "stem_cell_boost",
    "melphalan", "busulfan", "conditioning"
}

# -----------------------------------------------------------------------------
# Rule Engine Helpers
# -----------------------------------------------------------------------------

def normalize_text(value) -> str:
    """removes extra whitespace and normalizes text strings"""
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_category(value) -> str:
    """Cleans text and removes "(Not in provided Fiona's category list)" suffix"""
    text = normalize_text(value)
    return text.replace(" (Not in provided Fiona's category list)", "").strip()


def clean_token(value: str) -> str:
    """Lowercases, strips brackets/quotes, and normalizes individual drug tokens"""
    token = normalize_text(value).lower()
    token = token.strip().strip("[]'").strip('"')
    token = token.replace("[", "").replace("]", "")
    token = re.sub(r"\s+", " ", token).strip()
    return token


def split_list_string(value) -> list[str]:
    """Split values like 'a, b, c' or 'A + B + C' into clean lowercase tokens."""
    text = normalize_text(value)
    if not text:
        return []
    text = text.replace(" (Not in provided Fiona's category list)", "")
    raw_parts = text.split(" + ") if " + " in text else text.split(",")
    parts = []
    for part in raw_parts:
        token = clean_token(part)
        if token and token not in parts:
            parts.append(token)
    return parts


def normalize_regimen_string(value) -> str:
    """Normalize the full regimen string while preserving bracket/chunk structure."""
    text = normalize_text(value).lower()
    if not text:
        return ""
    text = text.replace("[ ", "[").replace(" ]", "]")
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\]\s*,\s*\[", "], [", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_regimen_chunks(value) -> list[tuple[str, ...]]:
    """Parse bracket-level regimen chunks."""
    text = normalize_regimen_string(value)
    if not text:
        return []

    bracket_chunks = re.findall(r"\[([^\]]+)\]", text)
    if not bracket_chunks:
        bracket_chunks = [text.replace("[", "").replace("]", "")]

    chunks: list[tuple[str, ...]] = []
    for chunk in bracket_chunks:
        drugs = []
        for raw in chunk.split(","):
            drug = clean_token(raw)
            if drug and drug not in drugs:
                drugs.append(drug)
        if drugs:
            chunks.append(tuple(drugs))
    return chunks


def chunk_list_to_str(chunks: list[tuple[str, ...]]) -> str:
    return " | ".join("[" + ", ".join(chunk) + "]" for chunk in chunks)


def get_drug_set_from_chunks(chunks: list[tuple[str, ...]]) -> set[str]:
    return {drug for chunk in chunks for drug in chunk}


def get_drug_set(row: pd.Series) -> set[str]:
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    if chunks:
        return get_drug_set_from_chunks(chunks)
    if "parsed_drugs" in row and normalize_text(row.get("parsed_drugs")):
        return set(split_list_string(row.get("parsed_drugs")))
    return set()


def get_active_drugs(row: pd.Series) -> set[str]:
    """Get drugs_active from pre-processing."""
    if "drugs_active" in row:
        active = row.get("drugs_active")
        if isinstance(active, str) and active:
            return set(split_list_string(active))
        elif isinstance(active, list):
            return set(active)
    # Fallback: parse from regimen and filter out steroids/conditioning
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    drugs = get_drug_set_from_chunks(chunks)
    steroids = {"dexamethasone", "prednisone", "prednisolone", "methylprednisolone"}
    conditioning = {"melphalan", "busulfan", "carmustine"}
    return {d for d in drugs if d not in steroids and d not in conditioning}


def get_tail_active_drugs(row: pd.Series) -> set[str]:
    """Get tail_drugs_active from pre-processing."""
    if "tail_drugs_active" in row:
        tail = row.get("tail_drugs_active")
        if isinstance(tail, str) and tail:
            return set(split_list_string(tail))
        elif isinstance(tail, list):
            return set(tail)
    return get_active_drugs(row)


def get_family_set(row: pd.Series) -> set[str]:
    if "mapped_family_classes" in row and normalize_text(row.get("mapped_family_classes")):
        return set(split_list_string(row.get("mapped_family_classes")))
    return set(split_list_string(clean_category(row.get("reconstructed_family_combination"))))


def has_complex_therapy(drug_set: set[str], family_set: set[str]) -> bool:
    combined = " ".join(sorted(drug_set | family_set)).lower()
    return any(keyword in combined for keyword in COMPLEX_THERAPY_KEYWORDS)


def set_to_str(values: set[str]) -> str:
    return ", ".join(sorted(values))


def get_regimen_signature(row: pd.Series) -> dict:
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    return {
        "normalized_regimen": normalize_regimen_string(row.get("line_of_therapy_name")),
        "chunks": chunks,
        "chunk_sets": [frozenset(chunk) for chunk in chunks],
        "drug_set": get_drug_set(row),
        "active_drugs": get_active_drugs(row),
        "tail_active_drugs": get_tail_active_drugs(row),
        "family_set": get_family_set(row),
    }


def calculate_gap_days(date1, date2) -> float:
    """Calculate gap in days between two dates."""
    if pd.isna(date1) or pd.isna(date2):
        return np.nan
    try:
        date1 = pd.to_datetime(date1)
        date2 = pd.to_datetime(date2)
        return float((date2 - date1).days)
    except:
        return np.nan

# -----------------------------------------------------------------------------
# Rule Engine Implementation
# -----------------------------------------------------------------------------

def rule_p1_confirmed_progression(row: pd.Series, prev_row: pd.Series, gap_days: float) -> tuple[bool, str]:
    """
    P1: Confirmed Progression → New LOT

    Trigger: A PD/Refractory response is recorded within the 60-day window
             before the start of the next segment.

    Exceptions (P1 does NOT fire):
    - The first maintenance segment immediately following ASCT
    - Same active drugs with gap ≤ 7 days (prescription renewal)
    - Active drugs are identical and gap ≤ 180 days → defer to Rules 2+3
    """
    # Check for PD signal
    discontinue_reason = normalize_text(row.get("discontinue_reason", ""))
    has_pd = any(term in discontinue_reason.lower() for term in ["pd", "progression", "refractory", "progressed"])

    if not has_pd:
        return False, ""

    # Check if first maintenance after ASCT (handled by ASCT Rule)
    if prev_row.get("is_asct", False):
        return False, "post-ASCT maintenance handled by ASCT Rule"

    # Check gap ≤ 7 days with same active drugs
    prev_active = get_active_drugs(prev_row)
    curr_active = get_active_drugs(row)
    if not pd.isna(gap_days) and gap_days <= 7 and prev_active == curr_active:
        return False, "prescription renewal (gap ≤ 7 days, same active drugs)"

    # Check if active drugs identical and gap ≤ 180 days → defer to Rules 2+3
    if not pd.isna(gap_days) and gap_days <= 180 and prev_active == curr_active:
        return False, "defer to Rules 2+3"

    return True, "PD/Refractory response within 60 days before segment start"


def rule_p1b_planned_sequential(row: pd.Series, prev_row: pd.Series, gap_days: float) -> tuple[bool, str]:
    """
    P1b: Planned Sequential Therapy → Same LOT

    All conditions must be met:
    1. Gap ≤ 3 days
    2. Next segment is single-agent LEN or THAL
    3. Prior segment contains ≥ 2 drugs in drugs_active (steroids excluded)
    4. Prior segment lasted 60-365 days
    5. No PD signal before the next segment starts
    """
    # Check gap
    if pd.isna(gap_days) or gap_days > 3:
        return False, "gap > 3 days"

    # Check current segment is single-agent LEN or THAL
    curr_active = get_active_drugs(row)
    if len(curr_active) != 1:
        return False, f"not single-agent: {len(curr_active)} active drugs"

    curr_drug = list(curr_active)[0]
    if curr_drug not in ["lenalidomide", "thalidomide"]:
        return False, f"not LEN or THAL: {curr_drug}"

    # Check prior segment has ≥ 2 active drugs
    prev_active = get_active_drugs(prev_row)
    if len(prev_active) < 2:
        return False, f"prior segment has {len(prev_active)} active drugs (< 2)"

    # Check prior segment duration
    if not pd.isna(prev_row.get("segment_duration_days")):
        duration = prev_row.get("segment_duration_days")
        if duration < 60 or duration > 365:
            return False, f"prior segment duration {duration} days (not 60-365)"

    # Check for PD signal
    discontinue_reason = normalize_text(row.get("discontinue_reason", ""))
    has_pd = any(term in discontinue_reason.lower() for term in ["pd", "progression", "refractory"])
    if has_pd:
        return False, "PD signal present"

    return True, "planned induction → maintenance sequence (e.g., VRd → LEN)"


def rule_steroid_only_absorbed(row: pd.Series, prev_row: pd.Series) -> tuple[bool, str]:
    """
    Rule 1: Steroid-Only Segment → Absorbed into Prior LOT

    Trigger: A segment contains only steroid drugs (e.g., dexamethasone alone)
             and the prior segment is not CAR-T.
    """
    is_steroid_only = row.get("is_steroid_only", False)
    if not is_steroid_only:
        return False, "not steroid-only segment"

    prev_is_car_t = prev_row.get("is_car_t", False)
    if prev_is_car_t:
        return False, "prior segment is CAR-T"

    return True, "steroid-only segment absorbed into prior LOT"


def rule_mandatory_drug_planned_triplet(row: pd.Series, prev_row: pd.Series,
                                        lot_start_date: any, gap_from_lot_start: float) -> tuple[bool, str]:
    """
    Mandatory-Drug Planned Triplet

    All conditions must be met:
    1. The next segment adds only a drug from the mandatory class:
       daratumumab, isatuximab, carfilzomib, or ixazomib
    2. It is a strict escalation (no drugs are removed from the prior segment)
    3. Addition occurs within 45 days of the first segment's start date
    """
    # Get active drugs
    prev_active = get_active_drugs(prev_row)
    curr_active = get_active_drugs(row)

    # Check: no drugs removed
    if not prev_active.issubset(curr_active):
        return False, "drugs removed from prior segment"

    # Check: only one drug added
    added_drugs = curr_active - prev_active
    if len(added_drugs) != 1:
        return False, f"{len(added_drugs)} drugs added"

    added_drug = list(added_drugs)[0]
    mandatory_drugs = {"daratumumab", "isatuximab", "carfilzomib", "ixazomib"}

    if added_drug not in mandatory_drugs:
        return False, f"added drug {added_drug} not in mandatory class"

    # Check: addition within 45 days of first segment start
    if pd.isna(gap_from_lot_start) or gap_from_lot_start > 45:
        return False, f"addition at day {gap_from_lot_start} > 45 days"

    return True, f"planned triplet: added {added_drug} within 45 days"


def rule_rules_2_3_identical_active_drugs(row: pd.Series, prev_row: pd.Series,
                                          gap_days: float) -> tuple[bool, str]:
    """
    Rules 2+3: Identical Active Drugs → Same LOT

    Trigger: tail_drugs_active is identical between segments AND gap ≤ 180 days.

    Special case: LEN single-agent → Rd (LEN + DEX):
    steroid re-escalation is flagged as ambiguous and reviewed by LLM.
    """
    # Check: tail_active_drugs identical
    prev_tail = get_tail_active_drugs(prev_row)
    curr_tail = get_tail_active_drugs(row)

    if prev_tail != curr_tail:
        return False, "tail_active_drugs not identical"

    # Check: gap ≤ 180 days
    if pd.isna(gap_days) or gap_days > 180:
        return False, f"gap {gap_days} days > 180"

    # Check if prior segment contains ASCT (exception)
    if prev_row.get("is_asct", False):
        return False, "prior segment contains ASCT (handled by ASCT rule)"

    # Special case: LEN single-agent → Rd
    if len(prev_tail) == 1 and "lenalidomide" in prev_tail:
        # Check if next is LEN + DEX (steroid re-escalation)
        next_active = get_active_drugs(row)
        if next_active == {"lenalidomide", "dexamethasone"}:
            # Flag for LLM review
            return True, "LEN → Rd (steroid re-escalation) - LLM review needed"

    return True, "identical active drugs with gap ≤ 180 days"


def rule_pre_asct_reinduction(row: pd.Series, prev_row: pd.Series,
                              prev_prev_row: pd.Series | None) -> tuple[bool, str]:
    """
    Pre-ASCT Re-induction

    All conditions must be true:
    1. Current segment contains ASCT
    2. The segment immediately preceding ASCT (re-induction) has drugs_active
       that is a proper subset of the segment before that (induction)
    3. Gap between re-induction start and ASCT ≤ 90 days
    """
    # Check current is ASCT
    if not row.get("is_asct", False):
        return False, "current segment not ASCT"

    # Check that we have both previous segments
    if prev_row is None or prev_prev_row is None:
        return False, "missing previous segments"

    # Check: re-induction active drugs is proper subset of induction
    prev_active = get_active_drugs(prev_row)  # re-induction
    prev_prev_active = get_active_drugs(prev_prev_row)  # induction

    if not prev_active.issubset(prev_prev_active):
        return False, "re-induction drugs not subset of induction drugs"

    if prev_active == prev_prev_active:
        return False, "re-induction drugs identical to induction (not proper subset)"

    # Check: gap between re-induction start and ASCT ≤ 90 days
    gap = calculate_gap_days(
        prev_row.get("date_start_line_of_therapy"),
        row.get("date_start_line_of_therapy")
    )

    if pd.isna(gap) or gap > 90:
        return False, f"gap {gap} days > 90"

    return True, "pre-ASCT re-induction (drugs dropped before transplant)"


def rule_asct_rule(row: pd.Series, prev_row: pd.Series,
                   asct_index: float) -> tuple[bool, str]:
    """
    ASCT Rule: Post-Transplant Maintenance

    Uses asct_index assigned during pre-processing.
    """
    if not row.get("is_asct", False) and not prev_row.get("is_asct", False):
        return False, "not ASCT-related transition"

    # Patient's first ASCT (asct_index = 1)
    if asct_index == 1:
        # Check if post-ASCT segment within 180 days
        gap = calculate_gap_days(
            prev_row.get("date_end_line_of_therapy"),
            row.get("date_start_line_of_therapy")
        )
        if not pd.isna(gap) and gap <= 180:
            return True, "first ASCT post-transplant maintenance (≤ 180 days)"

    # Patient's subsequent ASCT (asct_index ≥ 2)
    elif asct_index >= 2 and prev_row.get("asct_index", 0) < asct_index:
        # Check: post-ASCT drugs_active is a proper subset of pre-ASCT induction
        prev_active = get_active_drugs(prev_row)
        curr_active = get_active_drugs(row)

        if curr_active.issubset(prev_active) and curr_active != prev_active:
            return True, "subsequent ASCT maintenance (proper subset of induction)"

        # New drug not in pre-ASCT induction appears post-ASCT
        # Flag for LLM review
        if not curr_active.issubset(prev_active):
            return True, "new active drug post-ASCT - LLM review needed"

    return False, "no ASCT rule conditions met"


def rule_car_t_rule(row: pd.Series, prev_row: pd.Series) -> tuple[bool, str]:
    """
    CAR-T Rule
    """
    if not row.get("is_car_t", False) and not prev_row.get("is_car_t", False):
        return False, "not CAR-T transition"

    # Conditioning merge: prior segment started within 14 days of CAR-T infusion
    if prev_row.get("is_car_t", False):
        gap = calculate_gap_days(
            prev_row.get("date_start_line_of_therapy"),
            row.get("date_start_line_of_therapy")
        )
        if not pd.isna(gap) and gap <= 14:
            return True, "conditioning merge (≤ 14 days before CAR-T)"

    # Post-CAR-T exclusion: next segment starts within 30 days of CAR-T infusion
    if row.get("is_car_t", False):
        gap = calculate_gap_days(
            prev_row.get("date_end_line_of_therapy"),
            row.get("date_start_line_of_therapy")
        )
        if not pd.isna(gap) and gap <= 30:
            return True, "post-CAR-T exclusion (≤ 30 days after CAR-T)"

    return False, "CAR-T - needs LLM review"


def rule_p3_maintenance_after_combination(row: pd.Series, prev_row: pd.Series,
                                          gap_days: float) -> tuple[bool, str]:
    """
    P3: Maintenance After Combination → Same LOT

    All conditions must be met:
    1. Active drugs in next segment are a proper subset of prior segment
    2. Next segment is a single maintenance drug: LEN, THAL, DARA, or IXA
       (or bortezomib single-agent with specific criteria)
    3. That drug was present in prior segment's tail_drugs_active
    4. Prior segment was a combination (≥ 2 active drugs) lasting ≥ 30 days
    5. Gap criteria
    """
    # Check: active drugs are proper subset
    prev_active = get_active_drugs(prev_row)
    curr_active = get_active_drugs(row)

    if not curr_active.issubset(prev_active) or curr_active == prev_active:
        return False, "current active drugs not proper subset of prior"

    # Check: next segment is single maintenance drug
    if len(curr_active) != 1:
        return False, f"not single-agent: {len(curr_active)} active drugs"

    maintenance_drugs = {"lenalidomide", "thalidomide", "daratumumab", "ixazomib"}
    curr_drug = list(curr_active)[0]

    # Check for bortezomib special case
    is_bortezomib = curr_drug == "bortezomib"

    if not is_bortezomib and curr_drug not in maintenance_drugs:
        return False, f"not a maintenance drug: {curr_drug}"

    # Check: drug was present in prior tail_drugs_active
    prev_tail = get_tail_active_drugs(prev_row)
    if curr_drug not in prev_tail:
        return False, f"{curr_drug} not in prior tail_active_drugs"

    # Check: prior segment was combination (≥ 2 active drugs) lasting ≥ 30 days
    if len(prev_active) < 2:
        return False, "prior segment not a combination"

    # Check prior segment duration
    if not pd.isna(prev_row.get("segment_duration_days")):
        duration = prev_row.get("segment_duration_days")
        if duration < 30:
            return False, f"prior segment duration {duration} days (< 30)"

    # Check gap criteria
    if is_bortezomib:
        # Bortezomib-specific criteria
        prev_is_asct = prev_row.get("is_asct", False)
        if prev_is_asct:
            # Post-ASCT: gap ≤ 30 days, DEX allowed
            if pd.isna(gap_days) or gap_days <= 30:
                return True, "post-ASCT bortezomib maintenance (≤ 30 days, DEX allowed)"
        else:
            # Non-ASCT: gap ≤ 14 days, no DEX
            if not pd.isna(gap_days) and gap_days <= 14:
                # Check for DEX
                curr_drugs = get_drug_set(row)
                if "dexamethasone" not in curr_drugs:
                    return True, "non-ASCT bortezomib maintenance (≤ 14 days, no DEX)"
                else:
                    return False, "non-ASCT bortezomib with DEX (not allowed)"
    else:
        # LEN/THAL/DARA/IXA: gap ≤ 30 days
        if not pd.isna(gap_days) and gap_days <= 30:
            return True, f"maintenance {curr_drug} (gap ≤ 30 days)"

    return False, "gap criteria not met"


def rule_default_merge(row: pd.Series, prev_row: pd.Series) -> tuple[bool, str]:
    """Default: Merge (same LOT) if no rule above fires."""
    return True, "default merge (no specific rule matched)"


# -----------------------------------------------------------------------------
# Main Classification Function with Rule Engine
# -----------------------------------------------------------------------------

def classify_row_with_rules(row: pd.Series, prev_row: pd.Series | None,
                            earlier_same_doctor_rows: pd.DataFrame,
                            lot_start_dates: dict) -> dict:
    """Classify row using the rule engine from the specification."""

    # Get signatures
    cur_sig = get_regimen_signature(row)
    current_drugs = cur_sig["drug_set"]
    current_active = cur_sig["active_drugs"]
    current_chunks = cur_sig["chunks"]
    current_chunk_sets = cur_sig["chunk_sets"]
    current_families = cur_sig["family_set"]

    prev_sig = {"normalized_regimen": "", "chunks": [], "chunk_sets": [],
                "drug_set": set(), "active_drugs": set(),
                "tail_active_drugs": set(), "family_set": set()}
    previous_drugs = set()
    previous_active = set()
    previous_chunks = []
    previous_chunk_sets = []
    previous_families = set()

    if prev_row is not None:
        prev_sig = get_regimen_signature(prev_row)
        previous_drugs = prev_sig["drug_set"]
        previous_active = prev_sig["active_drugs"]
        previous_chunks = prev_sig["chunks"]
        previous_chunk_sets = prev_sig["chunk_sets"]
        previous_families = prev_sig["family_set"]

    added_drugs = current_drugs - previous_drugs
    removed_drugs = previous_drugs - current_drugs
    added_families = current_families - previous_families
    removed_families = previous_families - current_families

    # Calculate gaps
    gap_days = calculate_gap_days(
        prev_row.get("date_end_line_of_therapy") if prev_row is not None else None,
        row.get("date_start_line_of_therapy")
    )

    # Get ASCT index
    asct_index = row.get("asct_index", np.nan)

    # Look backward inside the same doctor LoT
    seen_same_full_regimen_earlier = False
    seen_same_chunk_structure_earlier = False
    seen_same_drug_set_earlier = False
    seen_same_family_set_earlier = False

    if not earlier_same_doctor_rows.empty:
        for _, earlier in earlier_same_doctor_rows.iterrows():
            earlier_sig = get_regimen_signature(earlier)
            if earlier_sig["normalized_regimen"] == cur_sig["normalized_regimen"] and cur_sig["normalized_regimen"]:
                seen_same_full_regimen_earlier = True
            if earlier_sig["chunks"] == current_chunks and current_chunks:
                seen_same_chunk_structure_earlier = True
            if earlier_sig["drug_set"] == current_drugs and current_drugs:
                seen_same_drug_set_earlier = True
            if earlier_sig["family_set"] == current_families and current_families:
                seen_same_family_set_earlier = True

    # Get lot start date for mandatory triplet rule
    lot_start_date = None
    gap_from_lot_start = np.nan
    if prev_row is not None:
        patient = row.get("cpid")
        if patient in lot_start_dates:
            lot_start_date = lot_start_dates[patient]
            gap_from_lot_start = calculate_gap_days(lot_start_date, row.get("date_start_line_of_therapy"))

    # Check for previous previous row (for pre-ASCT re-induction)
    prev_prev_row = None
    if prev_row is not None:
        # Find the row before prev_row
        patient_rows = earlier_same_doctor_rows[
            earlier_same_doctor_rows["_original_row_order"] < prev_row.get("_original_row_order", 0)]
        if not patient_rows.empty:
            prev_prev_row = patient_rows.iloc[-1]

    # ========================================================================
    # Apply Rule Engine - First Match Wins (Priority Order)
    # ========================================================================

    # If no previous row, this is the first patient row
    if prev_row is None or pd.isna(prev_row.get("cpid", "")):
        pattern = "unknown_pattern_requires_review"
        explanation = "First patient row - cannot compare to previous"
        rule_applied = "none"

    # P1: Confirmed Progression → New LOT
    else:
        p1_result, p1_explanation = rule_p1_confirmed_progression(row, prev_row, gap_days)
        if p1_result:
            pattern = "p1_confirmed_progression"
            explanation = f"P1: {p1_explanation}"
            rule_applied = "P1"

        # P1b: Planned Sequential Therapy → Same LOT
        elif rule_p1b_planned_sequential(row, prev_row, gap_days)[0]:
            pattern = "p1b_planned_sequential"
            explanation = "P1b: Planned induction → maintenance sequence (e.g., VRd → LEN)"
            rule_applied = "P1b"

        # Rule 1: Steroid-Only Segment → Absorbed into Prior LOT
        elif rule_steroid_only_absorbed(row, prev_row)[0]:
            pattern = "rule1_steroid_only_absorbed"
            explanation = "Rule 1: Steroid-only segment absorbed into prior LOT"
            rule_applied = "Rule 1"

        # Mandatory-Drug Planned Triplet
        elif not pd.isna(gap_from_lot_start) and \
                rule_mandatory_drug_planned_triplet(row, prev_row, lot_start_date, gap_from_lot_start)[0]:
            pattern = "mandatory_drug_planned_triplet"
            explanation = "Mandatory-drug planned triplet (added within 45 days of LOT start)"
            rule_applied = "Mandatory Triplet"

        # Rules 2+3: Identical Active Drugs → Same LOT
        elif rule_rules_2_3_identical_active_drugs(row, prev_row, gap_days)[0]:
            pattern = "rules_2_3_identical_active_drugs"
            explanation = "Rules 2+3: Identical active drugs with gap ≤ 180 days"
            rule_applied = "Rules 2+3"

        # Pre-ASCT Re-induction
        elif rule_pre_asct_reinduction(row, prev_row, prev_prev_row)[0]:
            pattern = "pre_asct_reinduction"
            explanation = "Pre-ASCT re-induction (drugs dropped before transplant)"
            rule_applied = "Pre-ASCT"

        # ASCT Rule: Post-Transplant Maintenance
        elif rule_asct_rule(row, prev_row, asct_index)[0]:
            pattern = "asct_rule_post_transplant"
            explanation = f"ASCT Rule (asct_index={asct_index}): Post-transplant maintenance"
            rule_applied = "ASCT Rule"

        # CAR-T Rule
        elif rule_car_t_rule(row, prev_row)[0]:
            pattern = "car_t_rule"
            explanation = "CAR-T Rule"
            rule_applied = "CAR-T Rule"

        # P3: Maintenance After Combination → Same LOT
        elif rule_p3_maintenance_after_combination(row, prev_row, gap_days)[0]:
            pattern = "p3_maintenance_after_combination"
            explanation = "P3: Maintenance after combination (de-escalation to single agent)"
            rule_applied = "P3"

        # Default: Merge (same LOT)
        else:
            pattern = "default_merge"
            explanation = "Default: Merge (same LOT) - no specific rule matched"
            rule_applied = "Default"

    # Determine review flag
    review_flag = 1 if pattern in FIXABLE_PATTERNS else 2
    suggested_corrected_lot = row.get("doctor_lot_numeric_for_transition") if review_flag == 1 else ""

    return {
        "misclassification_pattern": pattern,
        "pattern_explanation": explanation,
        "rule_applied": rule_applied,
        "review_flag": review_flag,
        "suggested_corrected_lot": suggested_corrected_lot,
        "previous_line_of_therapy_name": normalize_text(
            prev_row.get("line_of_therapy_name")) if prev_row is not None else "",
        "previous_reconstructed_family_combination": clean_category(
            prev_row.get("reconstructed_family_combination")) if prev_row is not None else "",
        "current_regimen_normalized": cur_sig["normalized_regimen"],
        "previous_regimen_normalized": prev_sig["normalized_regimen"],
        "current_regimen_chunks": chunk_list_to_str(current_chunks),
        "previous_regimen_chunks": chunk_list_to_str(previous_chunks),
        "current_drugs_clean": set_to_str(current_drugs),
        "previous_drugs_clean": set_to_str(previous_drugs),
        "current_active_drugs": set_to_str(current_active),
        "previous_active_drugs": set_to_str(previous_active),
        "added_drugs_vs_previous": set_to_str(added_drugs),
        "removed_drugs_vs_previous": set_to_str(removed_drugs),
        "current_families_clean": set_to_str(current_families),
        "previous_families_clean": set_to_str(previous_families),
        "added_families_vs_previous": set_to_str(added_families),
        "removed_families_vs_previous": set_to_str(removed_families),
        "gap_days_to_previous": gap_days,
    }


def count_lots_with_rule_engine(df_patient_rows):
    """
    Count Lines of Therapy using SCRIPT 1's rule engine.

    Parameters:
    -----------
    df_patient_rows : DataFrame
        All rows for a single patient, sorted chronologically

    Returns:
    --------
    dict with:
        - lot_assignments: List of (row_index, lot_number) for each row
        - total_lots: Total number of distinct LOTs for the patient
        - rule_applied: Which rule was used for each transition
        - flags: Any special flags (e.g., 'needs_review')
    """

    # Sort rows chronologically
    sorted_rows = df_patient_rows.sort_values('_original_row_order')

    # Initialize
    lot_assignments = []
    current_lot = 1
    rule_history = []
    flags = []

    # Start with first row
    first_row = sorted_rows.iloc[0]
    lot_assignments.append({
        'row_id': first_row['_original_row_order'],
        'lot_number': current_lot,
        'regimen': first_row.get('line_of_therapy_name', ''),
        'is_new_lot': False,
        'rule_applied': 'START'
    })

    # Process each subsequent row
    for idx in range(1, len(sorted_rows)):
        current_row = sorted_rows.iloc[idx]
        previous_row = sorted_rows.iloc[idx - 1]

        # Get earlier rows for look-back patterns
        earlier_rows = sorted_rows.iloc[:idx]

        # Build lot_start_dates for mandatory triplet rule
        lot_start_dates = {current_row['cpid']: sorted_rows.iloc[0].get('date_start_line_of_therapy')}

        # Use SCRIPT 1's classification function
        classification = classify_row_with_rules(
            current_row,
            previous_row,
            earlier_rows[
                earlier_rows['doctor_lot_numeric_for_transition'] == current_row['doctor_lot_numeric_for_transition']],
            lot_start_dates
        )

        pattern = classification['misclassification_pattern']
        rule_applied = classification['rule_applied']

        # Determine if this is a NEW LOT based on the pattern
        is_new_lot = determine_if_new_lot(pattern)

        if is_new_lot:
            current_lot += 1
            # Track why this was a new LOT
            if 'new_lot_triggers' not in locals():
                new_lot_triggers = []
            new_lot_triggers.append({
                'row': current_row['_original_row_order'],
                'reason': pattern,
                'rule': rule_applied
            })

        lot_assignments.append({
            'row_id': current_row['_original_row_order'],
            'lot_number': current_lot,
            'regimen': current_row.get('line_of_therapy_name', ''),
            'is_new_lot': is_new_lot,
            'rule_applied': rule_applied,
            'pattern': pattern,
            'explanation': classification['pattern_explanation']
        })

        # Check for flags that need human review
        if classification['review_flag'] == 2:
            flags.append({
                'row': current_row['_original_row_order'],
                'reason': pattern,
                'explanation': classification['pattern_explanation']
            })

    return {
        'total_lots': current_lot,
        'lot_assignments': lot_assignments,
        'rule_history': rule_history,
        'flags': flags,
        'new_lot_triggers': new_lot_triggers if 'new_lot_triggers' in locals() else []
    }


def determine_if_new_lot(pattern: str) -> bool:
    """
    Determine if a classification pattern indicates a NEW LOT.

    Based on the rule specification:
    - Some patterns clearly mean NEW LOT
    - Some patterns clearly mean MERGE (same LOT)
    - Some patterns are ambiguous (need review)
    """

    # Patterns that DEFINITELY mean NEW LOT
    new_lot_patterns = {
        'p1_confirmed_progression',  # PD = new line
        'p3_maintenance_after_combination_fallback',  # If P3 fails = new line
        'new_drug_addition_requires_review',  # Drug addition may be new line
        'complex_transplant_or_cell_therapy',  # Complex = often new line
        'complex_multi_drag_change_requires_review',  # Complex change
        'unknown_pattern_requires_review',  # Can't determine = new line
    }

    # Patterns that DEFINITELY mean MERGE (same LOT)
    merge_patterns = {
        'p1b_planned_sequential',  # VRd → LEN = same line
        'rule1_steroid_only_absorbed',  # Steroids only = not a line
        'mandatory_drug_planned_triplet',  # Planned escalation
        'rules_2_3_identical_active_drugs',  # Same drugs = same line
        'pre_asct_reinduction',  # Part of transplant plan
        'asct_rule_post_transplant',  # Post-transplant = same line
        'p3_maintenance_after_combination',  # Maintenance = same line
        'default_merge',  # Default = merge
        'same_exact_regimen_repeated',  # Same regimen = same line
        'same_drug_set_different_regimen_structure',  # Same drugs
        'regimen_phase_drop_or_deescalation',  # De-escalation
        'drug_drop_or_deescalation',  # Drug drop
        'family_drop_or_deescalation',  # Family drop
        're_expansion_to_prior_regimen',  # Re-expansion
        'same_family_drug_substitution',  # Family same = maybe same line
    }

    # Patterns that are AMBIGUOUS (default to new LOT for safety)
    ambiguous_patterns = {
        'same_family_drug_substitution',  # Family same but drug changed
    }

    if pattern in new_lot_patterns:
        return True
    elif pattern in merge_patterns:
        return False
    elif pattern in ambiguous_patterns:
        # For ambiguous cases, flag for review but default to new LOT
        # (better to over-count than under-count)
        return True
    else:
        # Unknown pattern - treat as new LOT
        return True


def get_lot_count_for_regimen(df_all_patients):
    """
    NEW FUNCTION ADDED:
    Get individual LoT counts for each regimen across all patients.
    """

    all_lot_data = []

    # Process each patient
    for patient in df_all_patients['cpid'].unique():
        patient_rows = df_all_patients[df_all_patients['cpid'] == patient]

        # Count lots for this patient
        result = count_lots_with_rule_engine(patient_rows)

        # Record each row's LOT assignment
        for assignment in result['lot_assignments']:
            all_lot_data.append({
                'cpid': patient,
                'row_id': assignment['row_id'],
                'lot_number': assignment['lot_number'],
                'total_lots_for_patient': result['total_lots'],
                'regimen': assignment['regimen'],
                'is_new_lot': assignment['is_new_lot'],
                'rule_applied': assignment.get('rule_applied', ''),
                'pattern': assignment.get('pattern', ''),
                'explanation': assignment.get('explanation', '')
            })

    lot_df = pd.DataFrame(all_lot_data)

    # Create regimen summary
    regimen_summary = (
        lot_df
        .groupby(['regimen', 'lot_number'])
        .agg({
            'cpid': 'nunique',
            'row_id': 'count'
        })
        .rename(columns={'cpid': 'patients_on_regimen', 'row_id': 'total_occurrences'})
        .reset_index()
        .sort_values(['regimen', 'lot_number'])
    )

    return lot_df, regimen_summary

# -----------------------------------------------------------------------------
# Load data
# -----------------------------------------------------------------------------
input_path = INPUT_PATH if INPUT_PATH.exists() else LEGACY_INPUT_PATH
if not input_path.exists():
    raise FileNotFoundError(
        f"No preprocessing output found. Expected {INPUT_PATH} "
        f"(run preprocessing_tester.py first) or {LEGACY_INPUT_PATH}."
    )

xl = pd.ExcelFile(input_path)

# preprocessing_tester.py writes "All_Data"; the legacy workbook writes
# "All_COTA_With_Transitions". Both hold every COTA row with transition columns.
all_rows_sheet = next(
    (name for name in ("All_Data", "All_COTA_With_Transitions") if name in xl.sheet_names),
    None,
)
if all_rows_sheet is None:
    raise ValueError(
        f"{input_path.name} must contain an 'All_Data' or "
        f"'All_COTA_With_Transitions' sheet. Found: {xl.sheet_names}"
    )

all_rows = pd.read_excel(input_path, sheet_name=all_rows_sheet)
all_rows.columns = all_rows.columns.str.strip()

# The legacy workbook ships a ready-made misclassified sheet. The newer one only
# has "Issues_Review", which also folds in needs_context rows, so derive the
# subset from the status column instead to keep the original row selection.
if "Misclassified_Rows" in xl.sheet_names:
    misclassified = pd.read_excel(input_path, sheet_name="Misclassified_Rows")
    misclassified.columns = misclassified.columns.str.strip()
else:
    if "transition_alignment_status" not in all_rows.columns:
        raise ValueError(
            f"Cannot derive misclassified rows: '{all_rows_sheet}' has no "
            "'transition_alignment_status' column."
        )
    misclassified = all_rows[
        all_rows["transition_alignment_status"].isin(MISCLASSIFIED_STATUSES)
    ].copy()

print(f"Input: {input_path.name} (all rows from '{all_rows_sheet}')")
print(f"Loaded {len(all_rows)} total rows, {len(misclassified)} misclassified rows")

required = ["cpid", "_original_row_order", "doctor_lot_numeric_for_transition"]
missing = [col for col in required if col not in all_rows.columns]
if missing:
    raise ValueError(f"Missing required columns in input workbook: {missing}")

all_rows = all_rows.sort_values(["cpid", "_original_row_order"]).copy()
misclassified = misclassified.sort_values(["cpid", "_original_row_order"]).copy()

all_rows["_row_key"] = all_rows["cpid"].astype(str) + "__" + all_rows["_original_row_order"].astype(str)
misclassified["_row_key"] = misclassified["cpid"].astype(str) + "__" + misclassified["_original_row_order"].astype(str)

# -----------------------------------------------------------------------------
# Build lot_start_dates dictionary for mandatory triplet rule
# -----------------------------------------------------------------------------
lot_start_dates = {}
for patient in all_rows["cpid"].unique():
    patient_rows = all_rows[all_rows["cpid"] == patient].sort_values("_original_row_order")
    if not patient_rows.empty:
        lot_start_dates[patient] = patient_rows.iloc[0].get("date_start_line_of_therapy")

# -----------------------------------------------------------------------------
# Classify each misclassified row with rule engine
# -----------------------------------------------------------------------------
classified_records = []

for _, row in misclassified.iterrows():
    patient_rows = all_rows[all_rows["cpid"].astype(str).eq(str(row["cpid"]))].copy()
    patient_rows = patient_rows.sort_values("_original_row_order")

    earlier_rows = patient_rows[patient_rows["_original_row_order"] < row["_original_row_order"]]
    prev_row = earlier_rows.iloc[-1] if not earlier_rows.empty else None

    same_doctor_lot = row.get("doctor_lot_numeric_for_transition")
    earlier_same_doctor_rows = earlier_rows[
        earlier_rows["doctor_lot_numeric_for_transition"].eq(same_doctor_lot)
    ]

    classified_records.append(classify_row_with_rules(row, prev_row, earlier_same_doctor_rows, lot_start_dates))

classification_df = pd.DataFrame(classified_records)
result = pd.concat([misclassified.reset_index(drop=True), classification_df], axis=1)

# -----------------------
# TRYING THE LOT COUNTER
# SEEMS TO BE HIGHLY INACCURATE...
# -----------------------

# After loading your data and applying the rule engine:
lot_df, regimen_summary = get_lot_count_for_regimen(all_rows)

# Output what you want:
print("\n=== LoT Counts by Regimen ===")
print(regimen_summary.to_string())

# Save to Excel with detailed results
with pd.ExcelWriter(LOT_COUNTS_PATH) as writer:
    lot_df.to_excel(writer, sheet_name='Detailed_LoT_Assignments', index=False)
    regimen_summary.to_excel(writer, sheet_name='Regimen_LoT_Summary', index=False)

    # Also create a pivot table
    pivot = regimen_summary.pivot(
        index='regimen',
        columns='lot_number',
        values='patients_on_regimen'
    ).fillna(0).astype(int)
    pivot.to_excel(writer, sheet_name='Regimen_LoT_Pivot')