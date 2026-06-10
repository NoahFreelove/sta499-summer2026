"""
Color-code COTA misclassified rows by broader misclassification case.

Input:
  Output/COTA_misclassified_rows.xlsx
  or COTA_misclassified_rows.xlsx in the same folder as this script.

Output:
  Output/COTA_misclassified_patterns_common_case_with_patient_counts.xlsx

Main update:
  These previously separate fixable subpatterns are grouped into ONE common case:
    - same_exact_regimen_repeated
    - same_drug_set_different_regimen_structure
    - regimen_phase_drop_or_deescalation
    - drug_drop_or_deescalation

The detailed subpattern is still preserved in a separate column.
  Case_Summary includes both row_count and patient_count.
"""

from __future__ import annotations

from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent

CANDIDATE_INPUTS = [
    BASE_DIR / "Output" / "COTA_misclassified_rows.xlsx",
    BASE_DIR / "Output" / "COTA_misclassified_rows_UPD.xlsx",
    BASE_DIR / "COTA_misclassified_rows.xlsx",
    BASE_DIR / "COTA_misclassified_rows_UPD.xlsx",
]

INPUT_PATH = next((p for p in CANDIDATE_INPUTS if p.exists()), None)
if INPUT_PATH is None:
    raise FileNotFoundError(
        "Could not find COTA_misclassified_rows.xlsx. Expected one of:\n"
        + "\n".join(str(p) for p in CANDIDATE_INPUTS)
    )

OUTPUT_DIR = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH = OUTPUT_DIR / "COTA_misclassified_patterns_common_case_with_patient_counts.xlsx"

COMMON_FIXABLE_SUBPATTERNS = {
    "same_exact_regimen_repeated",
    "same_drug_set_different_regimen_structure",
    "regimen_phase_drop_or_deescalation",
    "drug_drop_or_deescalation",
}

COMMON_FIXABLE_CASE = "regimen_continuation_or_deescalation_over_split"

CASE_COLORS = {
    COMMON_FIXABLE_CASE: "D9EAD3",
    "family_drop_or_deescalation": "FCE5CD",
    "re_expansion_to_prior_regimen": "D9EAF7",
    "new_drug_addition_requires_review": "F4CCCC",
    "same_family_drug_substitution": "EADCF8",
    "complex_transplant_or_cell_therapy": "D9D2E9",
    "complex_multi_drug_change_requires_review": "E6B8AF",
    "unknown_pattern_requires_review": "D9D9D9",
}

FIXABLE_CASES = {
    COMMON_FIXABLE_CASE,
    "family_drop_or_deescalation",
    "re_expansion_to_prior_regimen",
}

COMPLEX_THERAPY_KEYWORDS = {
    "transplant", "car-t", "cart", "autologous sct", "stem_cell_boost",
    "melphalan", "busulfan", "conditioning"
}


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_category(value) -> str:
    text = clean_text(value)
    return text.replace(" (Not in provided Fiona's category list)", "").strip()


def clean_token(value: str) -> str:
    token = clean_text(value).lower()
    token = token.strip().strip("[]'").strip('"')
    token = token.replace("[", "").replace("]", "")
    token = re.sub(r"\s+", " ", token).strip()
    return token


def split_list_string(value) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    text = text.replace(" (Not in provided Fiona's category list)", "")
    raw_parts = text.split(" + ") if " + " in text else text.split(",")
    parts = []
    for part in raw_parts:
        token = clean_token(part)
        if token and token not in parts:
            parts.append(token)
    return parts


def normalize_regimen_string(value) -> str:
    text = clean_text(value).lower()
    if not text:
        return ""
    text = text.replace("[ ", "[").replace(" ]", "]")
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\]\s*,\s*\[", "], [", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_regimen_chunks(value) -> list[tuple[str, ...]]:
    text = normalize_regimen_string(value)
    if not text:
        return []

    bracket_chunks = re.findall(r"\[([^\]]+)\]", text)
    if not bracket_chunks:
        bracket_chunks = [text.replace("[", "").replace("]", "")]

    chunks: list[tuple[str, ...]] = []
    for chunk in bracket_chunks:
        drugs = []
        for raw in chunk.split(","):
            drug = clean_token(raw)
            if drug and drug not in drugs:
                drugs.append(drug)
        if drugs:
            chunks.append(tuple(drugs))
    return chunks


def chunk_list_to_str(chunks: list[tuple[str, ...]]) -> str:
    return " | ".join("[" + ", ".join(chunk) + "]" for chunk in chunks)


def get_drug_set_from_chunks(chunks: list[tuple[str, ...]]) -> set[str]:
    return {drug for chunk in chunks for drug in chunk}


def get_regimen_signature(row: pd.Series) -> dict:
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    return {
        "normalized_regimen": normalize_regimen_string(row.get("line_of_therapy_name")),
        "chunks": chunks,
        "chunk_sets": [frozenset(chunk) for chunk in chunks],
        "drug_set": get_drug_set_from_chunks(chunks),
    }


def get_family_set(row: pd.Series) -> set[str]:
    if "mapped_family_classes" in row and clean_text(row.get("mapped_family_classes")):
        return set(split_list_string(row.get("mapped_family_classes")))
    return set(split_list_string(clean_category(row.get("reconstructed_family_combination"))))


def has_complex_therapy(drug_set: set[str], family_set: set[str]) -> bool:
    combined = " ".join(sorted(drug_set | family_set)).lower()
    return any(keyword in combined for keyword in COMPLEX_THERAPY_KEYWORDS)


def set_to_str(values: set[str]) -> str:
    return ", ".join(sorted(values))


def unique_join(values) -> str:
    seen = []
    for value in values:
        text = clean_text(value)
        if text and text not in seen:
            seen.append(text)
    return "; ".join(seen)


def chunk_sets_subset(current_chunk_sets: list[frozenset], previous_chunk_sets: list[frozenset]) -> bool:
    if not current_chunk_sets or not previous_chunk_sets:
        return False
    return all(chunk in previous_chunk_sets for chunk in current_chunk_sets)


def any_current_chunk_subset_of_previous(current_chunk_sets: list[frozenset], previous_chunk_sets: list[frozenset]) -> bool:
    if not current_chunk_sets or not previous_chunk_sets:
        return False
    for cur in current_chunk_sets:
        if cur in previous_chunk_sets:
            continue
        if any(cur.issubset(prev) for prev in previous_chunk_sets):
            return True
    return False


def convert_subpattern_to_case(subpattern: str) -> tuple[str, int, str, str]:
    """Return broad case, review flag, short meaning, fixability."""
    if subpattern in COMMON_FIXABLE_SUBPATTERNS:
        return (
            COMMON_FIXABLE_CASE,
            1,
            "COTA started a new LoT, but the change looks like continuation, repeat, or de-escalation within the same doctor LoT.",
            "Likely script-fixable: merge into the doctor LoT / previous aligned LoT."
        )
    if subpattern in FIXABLE_CASES:
        return (
            subpattern,
            1,
            "COTA over-split in a recurring pattern that can likely be detected by deterministic logic.",
            "Likely script-fixable, but review a sample before applying globally."
        )
    return (
        subpattern,
        2,
        "This change may be clinically meaningful or too structurally complex for a safe automatic correction.",
        "Human review needed."
    )


def classify_row(row: pd.Series, prev_row: pd.Series | None, earlier_same_doctor_rows: pd.DataFrame) -> dict:
    cur_sig = get_regimen_signature(row)
    current_drugs = cur_sig["drug_set"]
    current_chunks = cur_sig["chunks"]
    current_chunk_sets = cur_sig["chunk_sets"]
    current_families = get_family_set(row)

    prev_sig = {"normalized_regimen": "", "chunks": [], "chunk_sets": [], "drug_set": set()}
    previous_drugs = set()
    previous_chunks = []
    previous_chunk_sets = []
    previous_families = set()

    if prev_row is not None:
        prev_sig = get_regimen_signature(prev_row)
        previous_drugs = prev_sig["drug_set"]
        previous_chunks = prev_sig["chunks"]
        previous_chunk_sets = prev_sig["chunk_sets"]
        previous_families = get_family_set(prev_row)

    added_drugs = current_drugs - previous_drugs
    removed_drugs = previous_drugs - current_drugs
    added_families = current_families - previous_families
    removed_families = previous_families - current_families

    seen_same_full_regimen_earlier = False
    seen_same_chunk_structure_earlier = False
    seen_same_drug_set_earlier = False
    seen_same_family_set_earlier = False

    if not earlier_same_doctor_rows.empty:
        for _, earlier in earlier_same_doctor_rows.iterrows():
            earlier_sig = get_regimen_signature(earlier)
            if earlier_sig["normalized_regimen"] == cur_sig["normalized_regimen"] and cur_sig["normalized_regimen"]:
                seen_same_full_regimen_earlier = True
            if earlier_sig["chunks"] == current_chunks and current_chunks:
                seen_same_chunk_structure_earlier = True
            if earlier_sig["drug_set"] == current_drugs and current_drugs:
                seen_same_drug_set_earlier = True
            if get_family_set(earlier) == current_families and current_families:
                seen_same_family_set_earlier = True

    complex_case = has_complex_therapy(current_drugs, current_families) or has_complex_therapy(previous_drugs, previous_families)

    if not previous_drugs and not previous_families:
        subpattern = "unknown_pattern_requires_review"
        explanation = "Previous patient row was not found, so the change cannot be compared safely."

    elif cur_sig["normalized_regimen"] == prev_sig["normalized_regimen"] and cur_sig["normalized_regimen"]:
        subpattern = "same_exact_regimen_repeated"
        explanation = "Full normalized regimen string is identical to the previous patient row."

    elif current_chunks == previous_chunks and current_chunks:
        subpattern = "same_exact_regimen_repeated"
        explanation = "Bracket-level regimen structure is identical to the previous patient row."

    elif current_drugs == previous_drugs and current_chunks != previous_chunks:
        if chunk_sets_subset(current_chunk_sets, previous_chunk_sets):
            subpattern = "regimen_phase_drop_or_deescalation"
            explanation = "Unique drug set is the same, but one regimen phase/chunk was removed."
        elif any_current_chunk_subset_of_previous(current_chunk_sets, previous_chunk_sets):
            subpattern = "same_drug_set_different_regimen_structure"
            explanation = "Unique drug set is the same, but bracket-level regimen structure changed."
        else:
            subpattern = "same_drug_set_different_regimen_structure"
            explanation = "Unique drug set is identical, but the full regimen/chunk structure changed."

    elif (seen_same_full_regimen_earlier or seen_same_chunk_structure_earlier) and (added_drugs or removed_drugs):
        subpattern = "re_expansion_to_prior_regimen"
        explanation = "Exact regimen/chunk structure appeared earlier within the same doctor LoT."

    elif current_drugs and previous_drugs and current_drugs.issubset(previous_drugs):
        subpattern = "drug_drop_or_deescalation"
        explanation = "Current unique drug set is a subset of the previous unique drug set."

    elif current_families and previous_families and current_families.issubset(previous_families):
        subpattern = "family_drop_or_deescalation"
        explanation = "Current family combination is a subset of the previous family combination."

    elif complex_case:
        subpattern = "complex_transplant_or_cell_therapy"
        explanation = "Transition involves transplant/CAR-T/conditioning-related therapy."

    elif current_families == previous_families and current_drugs != previous_drugs:
        subpattern = "same_family_drug_substitution"
        explanation = "Drug names changed while the family combination stayed the same."

    elif added_drugs and not removed_drugs:
        if seen_same_full_regimen_earlier or seen_same_chunk_structure_earlier or seen_same_drug_set_earlier or seen_same_family_set_earlier:
            subpattern = "re_expansion_to_prior_regimen"
            explanation = "Regimen/drug/family combination reappears within the same doctor LoT."
        else:
            subpattern = "new_drug_addition_requires_review"
            explanation = "Drug was added; this can be clinically meaningful and needs review."

    elif added_drugs and removed_drugs:
        subpattern = "complex_multi_drug_change_requires_review"
        explanation = "Both added and removed drugs are present."

    else:
        subpattern = "unknown_pattern_requires_review"
        explanation = "No simple deterministic pattern was detected."

    broad_case, review_flag, broad_meaning, fixability = convert_subpattern_to_case(subpattern)
    suggested_corrected_lot = row.get("doctor_lot_numeric_for_transition") if review_flag == 1 else ""

    return {
        "misclassification_case": broad_case,
        "misclassification_subpattern": subpattern,
        "review_flag": review_flag,
        "suggested_corrected_lot": suggested_corrected_lot,
        "case_meaning": broad_meaning,
        "fixability_assessment": fixability,
        "subpattern_explanation": explanation,
        "previous_line_of_therapy_name": clean_text(prev_row.get("line_of_therapy_name")) if prev_row is not None else "",
        "previous_reconstructed_family_combination": clean_category(prev_row.get("reconstructed_family_combination")) if prev_row is not None else "",
        "current_regimen_normalized": cur_sig["normalized_regimen"],
        "previous_regimen_normalized": prev_sig["normalized_regimen"],
        "current_regimen_chunks": chunk_list_to_str(current_chunks),
        "previous_regimen_chunks": chunk_list_to_str(previous_chunks),
        "current_drugs_clean": set_to_str(current_drugs),
        "previous_drugs_clean": set_to_str(previous_drugs),
        "added_drugs_vs_previous": set_to_str(added_drugs),
        "removed_drugs_vs_previous": set_to_str(removed_drugs),
        "current_families_clean": set_to_str(current_families),
        "previous_families_clean": set_to_str(previous_families),
        "added_families_vs_previous": set_to_str(added_families),
        "removed_families_vs_previous": set_to_str(removed_families),
    }


xl = pd.ExcelFile(INPUT_PATH)
if "Misclassified_Rows" not in xl.sheet_names:
    raise ValueError("Input workbook must contain a 'Misclassified_Rows' sheet.")

misclassified = pd.read_excel(INPUT_PATH, sheet_name="Misclassified_Rows")
misclassified.columns = misclassified.columns.str.strip()

if "All_COTA_With_Transitions" in xl.sheet_names:
    all_rows = pd.read_excel(INPUT_PATH, sheet_name="All_COTA_With_Transitions")
    all_rows.columns = all_rows.columns.str.strip()
else:
    all_rows = misclassified.copy()

required = ["cpid", "_original_row_order", "doctor_lot_numeric_for_transition"]
missing = [col for col in required if col not in all_rows.columns]
if missing:
    raise ValueError(f"Missing required columns in input workbook: {missing}")

all_rows = all_rows.sort_values(["cpid", "_original_row_order"]).copy()
misclassified = misclassified.sort_values(["cpid", "_original_row_order"]).copy()

classified_records = []
for _, row in misclassified.iterrows():
    patient_rows = all_rows[all_rows["cpid"].astype(str).eq(str(row["cpid"]))].copy()
    patient_rows = patient_rows.sort_values("_original_row_order")

    earlier_rows = patient_rows[patient_rows["_original_row_order"] < row["_original_row_order"]]
    prev_row = earlier_rows.iloc[-1] if not earlier_rows.empty else None

    same_doctor_lot = row.get("doctor_lot_numeric_for_transition")
    earlier_same_doctor_rows = earlier_rows[
        earlier_rows["doctor_lot_numeric_for_transition"].eq(same_doctor_lot)
    ]

    classified_records.append(classify_row(row, prev_row, earlier_same_doctor_rows))

classification_df = pd.DataFrame(classified_records)
result = pd.concat([misclassified.reset_index(drop=True), classification_df], axis=1)

preferred_columns = [
    "misclassification_case",
    "misclassification_subpattern",
    "review_flag",
    "suggested_corrected_lot",
    "case_meaning",
    "fixability_assessment",
    "subpattern_explanation",
    "cpid",
    "line_of_therapy_c",
    "doctor_lot_numeric_for_transition",
    "previous_cota_lot_numeric",
    "previous_doctor_lot_numeric",
    "cota_minus_doctor_lot_shift",
    "reconstructed_family_combination",
    "line_of_therapy_name",
    "previous_line_of_therapy_name",
    "current_regimen_chunks",
    "previous_regimen_chunks",
    "current_regimen_normalized",
    "previous_regimen_normalized",
    "current_drugs_clean",
    "previous_drugs_clean",
    "added_drugs_vs_previous",
    "removed_drugs_vs_previous",
    "current_families_clean",
    "previous_families_clean",
    "added_families_vs_previous",
    "removed_families_vs_previous",
    "transition_alignment_status",
    "discontinue_reason",
    "date_start_line_of_therapy",
    "date_end_line_of_therapy",
]
preferred_columns = [col for col in preferred_columns if col in result.columns]
remaining_columns = [col for col in result.columns if col not in preferred_columns]
result = result[preferred_columns + remaining_columns]

case_summary = (
    result.groupby(["misclassification_case", "review_flag"], dropna=False)
    .agg(
        row_count=("cpid", "size"),
        patient_count=("cpid", "nunique"),
        patient_ids=("cpid", unique_join),
    )
    .reset_index()
    .sort_values(["row_count", "patient_count"], ascending=[False, False])
)

subpattern_summary = (
    result.groupby(["misclassification_case", "misclassification_subpattern", "review_flag"], dropna=False)
    .size()
    .reset_index(name="row_count")
    .sort_values(["misclassification_case", "row_count"], ascending=[True, False])
)

patient_summary = (
    result.groupby("cpid", dropna=False)
    .agg(
        total_misclassified_rows=("cpid", "size"),
        flag_1_fixable_rows=("review_flag", lambda x: int((x == 1).sum())),
        flag_2_human_review_rows=("review_flag", lambda x: int((x == 2).sum())),
        misclassified_cota_lots=("line_of_therapy_c", unique_join),
        doctor_lots=("doctor_lot_numeric_for_transition", unique_join),
        misclassification_cases=("misclassification_case", unique_join),
        misclassification_subpatterns=("misclassification_subpattern", unique_join),
        family_combinations=("reconstructed_family_combination", unique_join),
    )
    .reset_index()
    .sort_values(["total_misclassified_rows", "cpid"], ascending=[False, True])
)

patient_lots_cols = [
    "cpid",
    "line_of_therapy_c",
    "doctor_lot_numeric_for_transition",
    "misclassification_case",
    "misclassification_subpattern",
    "review_flag",
    "suggested_corrected_lot",
    "reconstructed_family_combination",
    "line_of_therapy_name",
    "previous_line_of_therapy_name",
    "case_meaning",
    "fixability_assessment",
    "subpattern_explanation",
]
patient_lots_cols = [col for col in patient_lots_cols if col in result.columns]
patient_lots = result[patient_lots_cols].sort_values(["cpid", "line_of_therapy_c"]).copy()

fixable = result[result["review_flag"].eq(1)].copy()
human_review = result[result["review_flag"].eq(2)].copy()

legend = pd.DataFrame([
    {
        "misclassification_case": COMMON_FIXABLE_CASE,
        "review_flag": 1,
        "color_hex": "#" + CASE_COLORS[COMMON_FIXABLE_CASE],
        "meaning": "Combined common fixable case: exact repeat, same-drug different structure, regimen phase drop/de-escalation, or drug drop/de-escalation.",
        "included_subpatterns": "; ".join(sorted(COMMON_FIXABLE_SUBPATTERNS)),
    },
    {
        "misclassification_case": "family_drop_or_deescalation",
        "review_flag": 1,
        "color_hex": "#" + CASE_COLORS["family_drop_or_deescalation"],
        "meaning": "Current family combination is a subset of the previous family combination.",
        "included_subpatterns": "family_drop_or_deescalation",
    },
    {
        "misclassification_case": "re_expansion_to_prior_regimen",
        "review_flag": 1,
        "color_hex": "#" + CASE_COLORS["re_expansion_to_prior_regimen"],
        "meaning": "Regimen/chunk/drug/family appeared earlier within the same doctor LoT.",
        "included_subpatterns": "re_expansion_to_prior_regimen",
    },
    {
        "misclassification_case": "new_drug_addition_requires_review",
        "review_flag": 2,
        "color_hex": "#" + CASE_COLORS["new_drug_addition_requires_review"],
        "meaning": "Drug addition may be clinically meaningful, so it needs review.",
        "included_subpatterns": "new_drug_addition_requires_review",
    },
    {
        "misclassification_case": "same_family_drug_substitution",
        "review_flag": 2,
        "color_hex": "#" + CASE_COLORS["same_family_drug_substitution"],
        "meaning": "Drug changed while family stayed the same; may still be clinically meaningful.",
        "included_subpatterns": "same_family_drug_substitution",
    },
    {
        "misclassification_case": "complex_transplant_or_cell_therapy",
        "review_flag": 2,
        "color_hex": "#" + CASE_COLORS["complex_transplant_or_cell_therapy"],
        "meaning": "Transplant/CAR-T/conditioning-related case; keep for clinician review.",
        "included_subpatterns": "complex_transplant_or_cell_therapy",
    },
    {
        "misclassification_case": "complex_multi_drug_change_requires_review",
        "review_flag": 2,
        "color_hex": "#" + CASE_COLORS["complex_multi_drug_change_requires_review"],
        "meaning": "Both additions and removals occurred; too complex for a safe automatic merge.",
        "included_subpatterns": "complex_multi_drug_change_requires_review",
    },
    {
        "misclassification_case": "unknown_pattern_requires_review",
        "review_flag": 2,
        "color_hex": "#" + CASE_COLORS["unknown_pattern_requires_review"],
        "meaning": "No deterministic pattern detected.",
        "included_subpatterns": "unknown_pattern_requires_review",
    },
])

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    result.to_excel(writer, sheet_name="Misclassified_Cases", index=False)
    case_summary.to_excel(writer, sheet_name="Case_Summary", index=False)
    subpattern_summary.to_excel(writer, sheet_name="Subpattern_Summary", index=False)
    patient_summary.to_excel(writer, sheet_name="Patient_Misclass_Summary", index=False)
    patient_lots.to_excel(writer, sheet_name="Patient_Misclass_LoTs", index=False)
    fixable.to_excel(writer, sheet_name="Flag_1_Fixable", index=False)
    human_review.to_excel(writer, sheet_name="Flag_2_Human_Review", index=False)
    legend.to_excel(writer, sheet_name="Legend", index=False)

wb = load_workbook(OUTPUT_PATH)
thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        header = str(ws.cell(row=1, column=col_idx).value or "")
        max_len = len(header)
        for cell in column_cells[1:80]:
            max_len = max(max_len, len(str(cell.value or "")))
        width = min(max(max_len + 2, 10), 45)
        if header in {
            "case_meaning", "fixability_assessment", "subpattern_explanation",
            "line_of_therapy_name", "previous_line_of_therapy_name",
            "reconstructed_family_combination", "current_regimen_chunks", "previous_regimen_chunks",
            "misclassification_cases", "misclassification_subpatterns", "family_combinations"
        }:
            width = 42
        ws.column_dimensions[get_column_letter(col_idx)].width = width

# Color-code row-level tabs by broad case.
for sheet_name in ["Misclassified_Cases", "Patient_Misclass_LoTs", "Flag_1_Fixable", "Flag_2_Human_Review"]:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    if "misclassification_case" not in headers:
        continue
    case_col_idx = headers.index("misclassification_case") + 1
    for row_idx in range(2, ws.max_row + 1):
        case = ws.cell(row=row_idx, column=case_col_idx).value
        fill_color = CASE_COLORS.get(case, CASE_COLORS["unknown_pattern_requires_review"])
        fill = PatternFill("solid", fgColor=fill_color)
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

# Color legend rows.
ws_legend = wb["Legend"]
legend_headers = [cell.value for cell in ws_legend[1]]
legend_case_col = legend_headers.index("misclassification_case") + 1
for row_idx in range(2, ws_legend.max_row + 1):
    case = ws_legend.cell(row=row_idx, column=legend_case_col).value
    fill_color = CASE_COLORS.get(case, CASE_COLORS["unknown_pattern_requires_review"])
    fill = PatternFill("solid", fgColor=fill_color)
    for col_idx in range(1, ws_legend.max_column + 1):
        ws_legend.cell(row=row_idx, column=col_idx).fill = fill

wb.save(OUTPUT_PATH)

print(f"Input: {INPUT_PATH}")
print(f"Saved: {OUTPUT_PATH}")
print(f"Total misclassified rows classified: {len(result)}")
print(f"Flag 1 likely script-fixable rows: {len(fixable)}")
print(f"Flag 2 human-review rows: {len(human_review)}")
print("\nCase counts:")
print(case_summary.to_string(index=False))
