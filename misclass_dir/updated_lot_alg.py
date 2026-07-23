"""
updated_lot_alg.py -- ATTEMPT 1
==========================
Combines the lot rule engine built from the updated rules (lot_rule_engine.py from Fiona) with COTA misclassification analysis.

This script:
1. Reads COTA data with transitions
2. Converts raw data into the Segment format expected by the rule engine
3. Runs the rule engine to determine LOTs
4. Compares rule engine output to COTA's LOT assignments
5. Classifies misclassifications by pattern
6. Outputs colored Excel with patterns and human review flags

Doesn't look in detail at previous files from Ana, e.g. reconstruct_cota...
Result... Seems similar to color_code_misclassification...
Output: unified_lot_adjudication_results
"""

from __future__ import annotations

from pathlib import Path
import re
from datetime import date, timedelta
from typing import Optional, Any
from dataclasses import dataclass, field
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
INPUT_PATH = BASE_DIR / "Output" / "COTA_misclassified_rows_UPD.xlsx"
OUTPUT_PATH = BASE_DIR / "Output" / "unified_lot_adjudication_results.xlsx"
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# -----------------------------------------------------------------------------
# PART 1: Rule Engine (from lot_rule_engine.py)
# -----------------------------------------------------------------------------

# Drug vocab (per rule_v1 "Definitions")
STEROIDS = {"dexamethasone", "prednisone", "prednisolone", "methylprednisolone"}
SCT_CONDITIONING_DRUGS = {"melphalan", "busulfan"}
SCT_TOKENS_RE = ("autologous", "allogeneic", "stem_cell_boost", "sct")
CART_TOKENS_RE = ("cart",)
MANDATORY_CLASS_DRUGS = {"daratumumab", "isatuximab", "carfilzomib", "ixazomib"}
LEN_THAL = {"lenalidomide", "thalidomide"}
P3_MAINTENANCE_DRUGS = {"lenalidomide", "thalidomide", "daratumumab", "ixazomib"}
BORTEZOMIB = "bortezomib"


def is_sct_token(token: str) -> bool:
    t = token.lower()
    return any(k in t for k in SCT_TOKENS_RE)


def is_cart_token(token: str) -> bool:
    t = token.lower()
    return any(k in t for k in CART_TOKENS_RE)


def active_drugs(tokens: set[str]) -> set[str]:
    """drugs_active: excludes steroids, SCT/conditioning drugs, and procedure markers."""
    out = set()
    for t in tokens:
        tl = t.lower().strip()
        if not tl or tl in STEROIDS or tl in SCT_CONDITIONING_DRUGS:
            continue
        if is_sct_token(tl) or is_cart_token(tl):
            continue
        out.add(tl)
    return out


@dataclass
class Segment:
    row_id: object
    start: Optional[date]
    end: Optional[date]
    chunks: list[tuple[str, ...]]
    pd_dates: list[date] = field(default_factory=list)
    contains_sct: bool = False
    contains_cart: bool = False
    asct_index: Optional[int] = None

    @property
    def all_tokens(self) -> set[str]:
        return {tok for chunk in self.chunks for tok in chunk}

    @property
    def drugs_active(self) -> set[str]:
        return active_drugs(self.all_tokens)

    @property
    def tail_drugs_active(self) -> set[str]:
        if not self.chunks:
            return set()
        return active_drugs(set(self.chunks[-1]))

    @property
    def is_steroid_only(self) -> bool:
        toks = self.all_tokens
        return bool(toks) and toks.issubset(STEROIDS)

    @property
    def n_active_drugs(self) -> int:
        return len(self.drugs_active)

    @property
    def duration_days(self) -> Optional[int]:
        if self.start and self.end:
            return (self.end - self.start).days
        return None


@dataclass
class Transition:
    outcome: str
    rule: str
    needs_llm_review: bool = False
    notes: str = ""


def gap_days(prev_end: Optional[date], next_start: Optional[date]) -> Optional[int]:
    if prev_end is None or next_start is None:
        return None
    return (next_start - prev_end).days


def pd_signal_in_window(pd_dates: list[date], window_start: Optional[date], window_end: Optional[date]) -> bool:
    """Any PD/refractory response date within (window_start, window_end]."""
    if window_end is None:
        return False
    lo = window_start if window_start is not None else date(1900, 1, 1)
    return any(lo < d <= window_end for d in pd_dates)


def absorb_steroid_only_first_segment(segments: list[Segment]) -> list[Segment]:
    if len(segments) < 2:
        return segments
    first, second = segments[0], segments[1]
    if first.is_steroid_only:
        g = gap_days(first.end, second.start)
        if g is not None and g <= 7:
            merged = Segment(
                row_id=second.row_id,
                start=first.start,
                end=second.end,
                chunks=first.chunks + second.chunks,
                pd_dates=second.pd_dates,
            )
            return [merged] + segments[2:]
    return segments


def assign_asct_index(segments: list[Segment]) -> None:
    idx = 0
    for seg in segments:
        seg.contains_sct = any(is_sct_token(tok) for tok in seg.all_tokens)
        seg.contains_cart = any(is_cart_token(tok) for tok in seg.all_tokens)
        if seg.contains_sct:
            idx += 1
            seg.asct_index = idx


def evaluate_transition(
        prev: Segment,
        cur: Segment,
        *,
        prior_prior: Optional[Segment],
        current_lot_first_start: Optional[date],
        current_lot_segments: list[Segment],
        prior_lot_induction: Optional[Segment],
) -> Transition:
    prev_active = prev.drugs_active
    cur_active = cur.drugs_active
    prev_tail = prev.tail_drugs_active
    cur_tail = cur.tail_drugs_active
    g = gap_days(prev.end, cur.start)

    # ---- P1: Confirmed Progression -> New LOT -----------------------------
    # FIXED: Check 60 days BEFORE cur.start, not just the gap
    pd_window_start = cur.start - timedelta(days=60) if cur.start else None
    pd_hit = pd_signal_in_window(cur.pd_dates, pd_window_start, cur.start) if cur.start else False

    if pd_hit:
        first_maint_after_first_asct = (prev.contains_sct and prev.asct_index == 1)
        same_drugs_short_gap = (prev_active == cur_active and g is not None and g <= 7)
        same_drugs_defer = (prev_active == cur_active and g is not None and g <= 180)

        if not (first_maint_after_first_asct or same_drugs_short_gap or same_drugs_defer):
            return Transition("new_lot", "P1_confirmed_progression",
                              notes=f"PD signal within 60-day pre-window.")

    # ---- P1b: Planned Sequential Therapy -> Same LOT -----------------------
    if (
            g is not None and g <= 3
            and len(cur_active) == 1 and next(iter(cur_active), None) in LEN_THAL
            and len(prev_active) >= 2
            and prev.duration_days is not None and 60 <= prev.duration_days <= 365
            and not pd_hit
    ):
        return Transition("merge", "P1b_planned_sequential_therapy",
                          notes="Induction -> single-agent LEN/THAL maintenance, planned sequence.")

    # ---- Rule 1: Steroid-Only Segment -> Absorbed --------------------------
    if cur.is_steroid_only and not prev.contains_cart:
        return Transition("merge", "R1_steroid_only_absorbed",
                          notes="Segment contains only steroids; absorbed into prior LOT.")

    # ---- Mandatory-Drug Planned Triplet ------------------------------------
    added = cur_active - prev_active
    removed = prev_active - cur_active
    if (
            added and added.issubset(MANDATORY_CLASS_DRUGS)
            and not removed
            and current_lot_first_start is not None and cur.start is not None
            and (cur.start - current_lot_first_start).days <= 45
    ):
        return Transition("merge", "mandatory_drug_planned_triplet",
                          notes=f"Added mandatory-class drug(s) {sorted(added)} within 45d of LOT start.")

    # ---- Rules 2+3: Identical Active Drugs (tail) -> Same LOT --------------
    skip_23 = prev.contains_sct
    if not skip_23 and prev_tail == cur_tail and prev_tail and g is not None and g <= 180:
        if prev_tail == {"lenalidomide"} and cur.all_tokens & STEROIDS and cur_tail == {"lenalidomide"}:
            if pd_hit:
                return Transition("new_lot", "R23_len_to_rd_pd_signal",
                                  needs_llm_review=False,
                                  notes="LEN->Rd re-escalation with PD signal present.")
            return Transition("merge", "R23_len_to_rd_ambiguous",
                              needs_llm_review=True,
                              notes="LEN->Rd steroid re-escalation: ambiguous, flagged for LLM/clinician review.")
        return Transition("merge", "R23_identical_active_drugs",
                          notes=f"tail_drugs_active identical ({sorted(prev_tail)}), gap={g}d <=180d.")

    # ---- Pre-ASCT Re-induction (looks back 2 segments) ---------------------
    if cur.contains_sct and prior_prior is not None:
        reinduction_subset_of_induction = (
                bool(prev.drugs_active) and bool(prior_prior.drugs_active)
                and prev.drugs_active < prior_prior.drugs_active
        )
        g_reinduction_to_asct = gap_days(prev.start, cur.start)
        if reinduction_subset_of_induction and g_reinduction_to_asct is not None and g_reinduction_to_asct <= 90:
            return Transition("merge", "pre_asct_reinduction",
                              notes="Drug(s) dropped immediately before transplant; induction+re-induction+ASCT merged.")

    # ---- ASCT Rule: Post-Transplant Maintenance ----------------------------
    if prev.contains_sct:
        if prev.asct_index == 1:
            days_since_asct = gap_days(prev.end, cur.start)
            if days_since_asct is not None and days_since_asct <= 180:
                return Transition("merge", "asct_first_maintenance",
                                  notes="First ASCT: post-transplant segment within 180d unconditionally merged.")
        elif prev.asct_index is not None and prev.asct_index >= 2:
            induction_active = prior_lot_induction.drugs_active if prior_lot_induction else set()
            days_since_asct = gap_days(prev.end, cur.start)
            within_window = days_since_asct is not None and days_since_asct <= 180
            if within_window and cur_active and induction_active and cur_active < induction_active:
                return Transition("merge", "asct_subsequent_maintenance_subset",
                                  notes="Subsequent ASCT: post-ASCT drugs are a proper subset of pre-ASCT induction.")
            if within_window and (cur_active - induction_active):
                return Transition("new_lot", "asct_subsequent_new_drug",
                                  needs_llm_review=True,
                                  notes=f"Subsequent ASCT: new drug(s) {sorted(cur_active - induction_active)} "
                                        "not in pre-ASCT induction regimen; flagged for LLM/clinician review.")

    # ---- CAR-T Rule ---------------------------------------------------------
    if cur.contains_cart or prev.contains_cart:
        if prev.contains_cart and not cur.contains_cart:
            days_after = gap_days(prev.end, cur.start)
            if days_after is not None and days_after <= 30:
                return Transition("merge", "cart_post_infusion_exclusion",
                                  notes="Segment starts within 30d of CAR-T infusion; merged into CAR-T LOT.")
            return Transition("new_lot", "cart_other_transition",
                              needs_llm_review=True,
                              notes="Post-CAR-T transition outside 30d window; flagged for LLM review.")
        if cur.contains_cart and not prev.contains_cart:
            g_cond = gap_days(prev.start, cur.start)
            if g_cond is not None and g_cond <= 14:
                return Transition("merge", "cart_conditioning_merge",
                                  notes="Prior segment started within 14d of CAR-T infusion; merged into CAR-T LOT.")
            return Transition("new_lot", "cart_other_transition",
                              needs_llm_review=True,
                              notes="Transition into CAR-T outside 14d conditioning window; flagged for LLM review.")

    # ---- P3: Maintenance After Combination -> Same LOT ---------------------
    is_single_maintenance_drug = len(cur_active) == 1 and next(iter(cur_active)) in P3_MAINTENANCE_DRUGS
    is_bortezomib_single = cur_active == {BORTEZOMIB}
    if (added == set() and removed and (is_single_maintenance_drug or is_bortezomib_single)
            and cur_active and cur_active < prev_active
            and next(iter(cur_active)) in prev_tail | (prev_active if not prev.chunks else set(prev.chunks[-1]))
            and len(prev_active) >= 2 and prev.duration_days is not None and prev.duration_days >= 30):
        if is_bortezomib_single:
            post_asct = prior_lot_induction is not None and any(s.contains_sct for s in current_lot_segments)
            if post_asct and g is not None and g <= 30:
                return Transition("merge", "p3_bortezomib_maintenance_post_asct",
                                  notes="Bortezomib single-agent maintenance post-ASCT, gap<=30d.")
            if (not post_asct) and g is not None and g <= 14 and not (cur.all_tokens & STEROIDS):
                return Transition("merge", "p3_bortezomib_maintenance_non_asct",
                                  notes="Bortezomib single-agent maintenance (non-ASCT), gap<=14d, no DEX.")
        else:
            if g is not None and g <= 30:
                return Transition("merge", "p3_single_maintenance_after_combo",
                                  notes=f"Single maintenance drug {sorted(cur_active)} continued from prior combo, gap<=30d.")
        return Transition("new_lot", "p3_conditions_not_met_drug_reduced",
                          notes="De-escalation looked like P3 maintenance but gap/context criteria failed.")

    # ---- Default -------------------------------------------------------------
    return Transition("merge", "default_merge", notes="No rule fired; default is merge (same LOT).")


def adjudicate_patient(raw_segments: list[Segment]) -> list[dict]:
    if not raw_segments:
        return []

    segments = absorb_steroid_only_first_segment(list(raw_segments))
    assign_asct_index(segments)

    results = []
    lot_num = 1
    current_lot_segments = [segments[0]]
    current_lot_first_start = segments[0].start
    prior_lot_induction = None

    results.append({
        "row_id": segments[0].row_id,
        "rule_lot": lot_num,
        "rule_fired": "line_zero_or_first_segment",
        "needs_llm_review": False,
        "notes": "First observed segment for this patient.",
    })

    for i in range(1, len(segments)):
        prev = segments[i - 1]
        cur = segments[i]
        prior_prior = segments[i - 2] if i >= 2 else None

        decision = evaluate_transition(
            prev, cur,
            prior_prior=prior_prior,
            current_lot_first_start=current_lot_first_start,
            current_lot_segments=current_lot_segments,
            prior_lot_induction=prior_lot_induction,
        )

        if decision.outcome == "new_lot":
            prior_lot_induction = current_lot_segments[0]
            lot_num += 1
            current_lot_segments = [cur]
            current_lot_first_start = cur.start
        else:
            current_lot_segments.append(cur)

        results.append({
            "row_id": cur.row_id,
            "rule_lot": lot_num,
            "rule_fired": decision.rule,
            "needs_llm_review": decision.needs_llm_review,
            "notes": decision.notes,
        })

    return results


# -----------------------------------------------------------------------------
# PART 2: COTA Misclassification Analysis (adapted from first script)
# -----------------------------------------------------------------------------

# Colors for misclassification patterns
PATTERN_COLORS = {
    "same_exact_regimen_repeated": "D9EAD3",
    "same_drug_set_different_regimen_structure": "B6D7A8",
    "regimen_phase_drop_or_deescalation": "CFE2F3",
    "drug_drop_or_deescalation": "FFF2CC",
    "family_drop_or_deescalation": "FCE5CD",
    "re_expansion_to_prior_regimen": "D9EAF7",
    "new_drug_addition_requires_review": "F4CCCC",
    "same_family_drug_substitution": "EADCF8",
    "complex_transplant_or_cell_therapy": "D9D2E9",
    "complex_multi_drug_change_requires_review": "E6B8AF",
    "unknown_pattern_requires_review": "D9D9D9",
    "rule_engine_agrees_with_cota": "E2EFDA",
    "rule_engine_disagrees_needs_review": "FFD9D9",
}

FIXABLE_PATTERNS = {
    "same_exact_regimen_repeated",
    "same_drug_set_different_regimen_structure",
    "regimen_phase_drop_or_deescalation",
    "drug_drop_or_deescalation",
    "family_drop_or_deescalation",
    "re_expansion_to_prior_regimen",
}

COMPLEX_THERAPY_KEYWORDS = {
    "transplant", "car-t", "cart", "autologous sct", "stem_cell_boost",
    "melphalan", "busulfan", "conditioning"
}


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_category(value) -> str:
    text = clean_text(value)
    return text.replace(" (Not in provided Fiona's category list)", "").strip()


def clean_token(value: str) -> str:
    token = clean_text(value).lower()
    token = token.strip().strip("[]'").strip('"')
    token = token.replace("[", "").replace("]", "")
    token = re.sub(r"\s+", " ", token).strip()
    return token


def split_list_string(value) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    text = text.replace(" (Not in provided Fiona's category list)", "")
    raw_parts = text.split(" + ") if " + " in text else text.split(",")
    parts = []
    for part in raw_parts:
        token = clean_token(part)
        if token and token not in parts:
            parts.append(token)
    return parts


def normalize_regimen_string(value) -> str:
    text = clean_text(value).lower()
    if not text:
        return ""
    text = text.replace("[ ", "[").replace(" ]", "]")
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\]\s*,\s*\[", "], [", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_regimen_chunks(value) -> list[tuple[str, ...]]:
    text = normalize_regimen_string(value)
    if not text:
        return []

    bracket_chunks = re.findall(r"\[([^\]]+)\]", text)
    if not bracket_chunks:
        bracket_chunks = [text.replace("[", "").replace("]", "")]

    chunks: list[tuple[str, ...]] = []
    for chunk in bracket_chunks:
        drugs = []
        for raw in chunk.split(","):
            drug = clean_token(raw)
            if drug and drug not in drugs:
                drugs.append(drug)
        if drugs:
            chunks.append(tuple(drugs))
    return chunks


def chunk_list_to_str(chunks: list[tuple[str, ...]]) -> str:
    return " | ".join("[" + ", ".join(chunk) + "]" for chunk in chunks)


def get_drug_set_from_chunks(chunks: list[tuple[str, ...]]) -> set[str]:
    return {drug for chunk in chunks for drug in chunk}


def get_drug_set(row: pd.Series) -> set[str]:
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    if chunks:
        return get_drug_set_from_chunks(chunks)
    if "parsed_drugs" in row and clean_text(row.get("parsed_drugs")):
        return set(split_list_string(row.get("parsed_drugs")))
    return set()


def get_regimen_signature(row: pd.Series) -> dict:
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    return {
        "normalized_regimen": normalize_regimen_string(row.get("line_of_therapy_name")),
        "chunks": chunks,
        "chunk_sets": [frozenset(chunk) for chunk in chunks],
        "drug_set": get_drug_set_from_chunks(chunks),
    }


def get_family_set(row: pd.Series) -> set[str]:
    if "mapped_family_classes" in row and clean_text(row.get("mapped_family_classes")):
        return set(split_list_string(row.get("mapped_family_classes")))
    return set(split_list_string(clean_category(row.get("reconstructed_family_combination"))))


def has_complex_therapy(drug_set: set[str], family_set: set[str]) -> bool:
    combined = " ".join(sorted(drug_set | family_set)).lower()
    return any(keyword in combined for keyword in COMPLEX_THERAPY_KEYWORDS)


def set_to_str(values: set[str]) -> str:
    return ", ".join(sorted(values))


def chunk_sets_subset(current_chunk_sets: list[frozenset], previous_chunk_sets: list[frozenset]) -> bool:
    if not current_chunk_sets or not previous_chunk_sets:
        return False
    return all(chunk in previous_chunk_sets for chunk in current_chunk_sets)


def any_current_chunk_subset_of_previous(current_chunk_sets: list[frozenset],
                                         previous_chunk_sets: list[frozenset]) -> bool:
    if not current_chunk_sets or not previous_chunk_sets:
        return False
    for cur in current_chunk_sets:
        if cur in previous_chunk_sets:
            continue
        if any(cur.issubset(prev) for prev in previous_chunk_sets):
            return True
    return False


def parse_date(value) -> Optional[date]:
    if pd.isna(value):
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    try:
        return pd.to_datetime(value).date()
    except:
        return None


def row_to_segment(row: pd.Series, pd_dates_by_patient: dict) -> Segment:
    """Convert a COTA row to a Segment object for the rule engine."""
    chunks = parse_regimen_chunks(row.get("line_of_therapy_name"))
    if not chunks:
        chunks = [tuple(split_list_string(row.get("parsed_drugs", "")))]

    start = parse_date(row.get("date_start_line_of_therapy"))
    end = parse_date(row.get("date_end_line_of_therapy"))

    cpid = str(row.get("cpid"))
    pd_dates = pd_dates_by_patient.get(cpid, [])

    return Segment(
        row_id=row.get("_row_key", row.get("_original_row_order")),
        start=start,
        end=end,
        chunks=chunks,
        pd_dates=pd_dates,
    )


def classify_misclassification(row: pd.Series, prev_row: pd.Series | None,
                               earlier_same_doctor_rows: pd.DataFrame) -> dict:
    cur_sig = get_regimen_signature(row)
    current_drugs = cur_sig["drug_set"]
    current_chunks = cur_sig["chunks"]
    current_chunk_sets = cur_sig["chunk_sets"]
    current_families = get_family_set(row)

    prev_sig = {"normalized_regimen": "", "chunks": [], "chunk_sets": [], "drug_set": set()}
    previous_drugs = set()
    previous_chunks = []
    previous_chunk_sets = []
    previous_families = set()

    if prev_row is not None:
        prev_sig = get_regimen_signature(prev_row)
        previous_drugs = prev_sig["drug_set"]
        previous_chunks = prev_sig["chunks"]
        previous_chunk_sets = prev_sig["chunk_sets"]
        previous_families = get_family_set(prev_row)

    added_drugs = current_drugs - previous_drugs
    removed_drugs = previous_drugs - current_drugs
    added_families = current_families - previous_families
    removed_families = previous_families - current_families

    seen_same_full_regimen_earlier = False
    seen_same_chunk_structure_earlier = False
    seen_same_drug_set_earlier = False
    seen_same_family_set_earlier = False

    if not earlier_same_doctor_rows.empty:
        for _, earlier in earlier_same_doctor_rows.iterrows():
            earlier_sig = get_regimen_signature(earlier)
            if earlier_sig["normalized_regimen"] == cur_sig["normalized_regimen"] and cur_sig["normalized_regimen"]:
                seen_same_full_regimen_earlier = True
            if earlier_sig["chunks"] == current_chunks and current_chunks:
                seen_same_chunk_structure_earlier = True
            if earlier_sig["drug_set"] == current_drugs and current_drugs:
                seen_same_drug_set_earlier = True
            if get_family_set(earlier) == current_families and current_families:
                seen_same_family_set_earlier = True

    complex_case = has_complex_therapy(current_drugs, current_families) or has_complex_therapy(previous_drugs,
                                                                                               previous_families)

    # Check if rule engine agrees with COTA
    rule_agrees = row.get("cota_lot_matches_rule", False)

    if rule_agrees:
        pattern = "rule_engine_agrees_with_cota"
        explanation = "Rule engine and COTA agree on LOT assignment."
        review_flag = 0  # No review needed
    elif not previous_drugs and not previous_families:
        pattern = "unknown_pattern_requires_review"
        explanation = "Previous patient row was not found, so the change cannot be compared safely."
        review_flag = 2
    elif cur_sig["normalized_regimen"] == prev_sig["normalized_regimen"] and cur_sig["normalized_regimen"]:
        pattern = "same_exact_regimen_repeated"
        explanation = "Full normalized regimen string is identical to previous patient row."
        review_flag = 1
    elif current_chunks == previous_chunks and current_chunks:
        pattern = "same_exact_regimen_repeated"
        explanation = "Bracket-level regimen structure is identical to previous patient row."
        review_flag = 1
    elif current_drugs == previous_drugs and current_chunks != previous_chunks:
        if chunk_sets_subset(current_chunk_sets, previous_chunk_sets):
            pattern = "regimen_phase_drop_or_deescalation"
            explanation = "Same drug set but one regimen phase/chunk was removed."
            review_flag = 1
        elif any_current_chunk_subset_of_previous(current_chunk_sets, previous_chunk_sets):
            pattern = "same_drug_set_different_regimen_structure"
            explanation = "Same drug set but bracket-level regimen structure changed."
            review_flag = 1
        else:
            pattern = "same_drug_set_different_regimen_structure"
            explanation = "Same drug set but regimen/chunk structure changed."
            review_flag = 1
    elif (seen_same_full_regimen_earlier or seen_same_chunk_structure_earlier) and (added_drugs or removed_drugs):
        pattern = "re_expansion_to_prior_regimen"
        explanation = "Regimen/chunk structure appeared earlier; likely shrink/re-expand cycling."
        review_flag = 1
    elif current_drugs and previous_drugs and current_drugs.issubset(previous_drugs):
        pattern = "drug_drop_or_deescalation"
        explanation = "Current unique drug set is a subset of previous; likely over-split after de-escalation."
        review_flag = 1
    elif current_families and previous_families and current_families.issubset(previous_families):
        pattern = "family_drop_or_deescalation"
        explanation = "Current family combination is a subset of previous; likely de-escalation."
        review_flag = 1
    elif complex_case:
        pattern = "complex_transplant_or_cell_therapy"
        explanation = "Transition involves transplant/CAR-T/conditioning-related therapy."
        review_flag = 2
    elif current_families == previous_families and current_drugs != previous_drugs:
        pattern = "same_family_drug_substitution"
        explanation = "Drug changed while family combination stayed same."
        review_flag = 2
    elif added_drugs and not removed_drugs:
        if seen_same_full_regimen_earlier or seen_same_chunk_structure_earlier or seen_same_drug_set_earlier or seen_same_family_set_earlier:
            pattern = "re_expansion_to_prior_regimen"
            explanation = "Regimen/drug/family combination reappeared; likely add-back/re-expansion."
            review_flag = 1
        else:
            pattern = "new_drug_addition_requires_review"
            explanation = "Drug added; may be clinically meaningful."
            review_flag = 2
    elif added_drugs and removed_drugs:
        pattern = "complex_multi_drug_change_requires_review"
        explanation = "Both additions and removals; complex swap/change."
        review_flag = 2
    else:
        pattern = "unknown_pattern_requires_review"
        explanation = "No simple deterministic pattern was detected."
        review_flag = 2

    return {
        "misclassification_pattern": pattern,
        "pattern_explanation": explanation,
        "review_flag": review_flag,
        "previous_line_of_therapy_name": clean_text(
            prev_row.get("line_of_therapy_name")) if prev_row is not None else "",
        "current_regimen_normalized": cur_sig["normalized_regimen"],
        "previous_regimen_normalized": prev_sig["normalized_regimen"],
        "current_regimen_chunks": chunk_list_to_str(current_chunks),
        "previous_regimen_chunks": chunk_list_to_str(previous_chunks),
        "current_drugs_clean": set_to_str(current_drugs),
        "previous_drugs_clean": set_to_str(previous_drugs),
        "added_drugs_vs_previous": set_to_str(added_drugs),
        "removed_drugs_vs_previous": set_to_str(removed_drugs),
        "current_families_clean": set_to_str(current_families),
        "previous_families_clean": set_to_str(previous_families),
        "added_families_vs_previous": set_to_str(added_families),
        "removed_families_vs_previous": set_to_str(removed_families),
    }


# -----------------------------------------------------------------------------
# Main execution
# -----------------------------------------------------------------------------

def main():
    print("Loading data...")
    xl = pd.ExcelFile(INPUT_PATH)

    if "Misclassified_Rows" not in xl.sheet_names:
        raise ValueError("Input workbook must contain a 'Misclassified_Rows' sheet.")

    misclassified = pd.read_excel(INPUT_PATH, sheet_name="Misclassified_Rows")
    misclassified.columns = misclassified.columns.str.strip()

    if "All_COTA_With_Transitions" in xl.sheet_names:
        all_rows = pd.read_excel(INPUT_PATH, sheet_name="All_COTA_With_Transitions")
        all_rows.columns = all_rows.columns.str.strip()
    else:
        all_rows = misclassified.copy()

    required = ["cpid", "_original_row_order", "doctor_lot_numeric_for_transition"]
    missing = [col for col in required if col not in all_rows.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    all_rows = all_rows.sort_values(["cpid", "_original_row_order"]).copy()
    misclassified = misclassified.sort_values(["cpid", "_original_row_order"]).copy()

    all_rows["_row_key"] = all_rows["cpid"].astype(str) + "__" + all_rows["_original_row_order"].astype(str)
    misclassified["_row_key"] = misclassified["cpid"].astype(str) + "__" + misclassified["_original_row_order"].astype(
        str)

    # Build PD dates per patient
    print("Building PD dates...")
    pd_dates_by_patient = {}
    for _, row in all_rows.iterrows():
        cpid = str(row.get("cpid"))
        if cpid not in pd_dates_by_patient:
            pd_dates_by_patient[cpid] = []

        discontinue_reason = clean_text(row.get("discontinue_reason", ""))
        if "pd" in discontinue_reason.lower() or "refractory" in discontinue_reason.lower():
            end_date = parse_date(row.get("date_end_line_of_therapy"))
            if end_date:
                pd_dates_by_patient[cpid].append(end_date)

    # Run rule engine on all patients
    print("Running rule engine...")
    all_results = []

    for cpid in all_rows["cpid"].unique():
        patient_rows = all_rows[all_rows["cpid"].astype(str) == str(cpid)].sort_values("_original_row_order")

        # Convert to segments
        segments = []
        for _, row in patient_rows.iterrows():
            seg = row_to_segment(row, pd_dates_by_patient)
            segments.append(seg)

        # Adjudicate
        results = adjudicate_patient(segments)
        all_results.extend(results)

    # Merge results back
    results_df = pd.DataFrame(all_results)
    all_rows_with_rules = all_rows.merge(
        results_df[["row_id", "rule_lot", "rule_fired", "needs_llm_review", "notes"]],
        left_on="_row_key",
        right_on="row_id",
        how="left"
    )

    # Flag where rule engine agrees/disagrees with COTA
    all_rows_with_rules["cota_lot_matches_rule"] = all_rows_with_rules["line_of_therapy_c"] == all_rows_with_rules[
        "rule_lot"]

    # Now classify misclassifications
    print("Classifying misclassifications...")
    classified_records = []

    for _, row in misclassified.iterrows():
        patient_rows = all_rows_with_rules[all_rows_with_rules["cpid"].astype(str) == str(row["cpid"])].copy()
        patient_rows = patient_rows.sort_values("_original_row_order")

        earlier_rows = patient_rows[patient_rows["_original_row_order"] < row["_original_row_order"]]
        prev_row = earlier_rows.iloc[-1] if not earlier_rows.empty else None

        same_doctor_lot = row.get("doctor_lot_numeric_for_transition")
        earlier_same_doctor_rows = earlier_rows[
            earlier_rows["doctor_lot_numeric_for_transition"].eq(same_doctor_lot)
        ]

        # Add rule agreement info to the row for classification
        row_copy = row.copy()
        if prev_row is not None:
            row_copy["cota_lot_matches_rule"] = row_copy.get("cota_lot_matches_rule", False)

        classified_records.append(classify_misclassification(row_copy, prev_row, earlier_same_doctor_rows))

    classification_df = pd.DataFrame(classified_records)
    result = pd.concat([misclassified.reset_index(drop=True), classification_df], axis=1)

    # Add rule engine results to the output
    result = result.merge(
        all_rows_with_rules[
            ["_row_key", "rule_lot", "rule_fired", "needs_llm_review", "notes", "cota_lot_matches_rule"]],
        left_on="_row_key",
        right_on="_row_key",
        how="left"
    )

    # Reorder columns
    preferred_columns = [
        "misclassification_pattern",
        "review_flag",
        "rule_lot",
        "rule_fired",
        "cota_lot_matches_rule",
        "needs_llm_review",
        "notes",
        "pattern_explanation",
        "cpid",
        "line_of_therapy_c",
        "doctor_lot_numeric_for_transition",
        "line_of_therapy_name",
        "current_regimen_chunks",
        "previous_regimen_chunks",
    ]
    preferred_columns = [col for col in preferred_columns if col in result.columns]
    remaining_columns = [col for col in result.columns if col not in preferred_columns]
    result = result[preferred_columns + remaining_columns]

    # Create summary tables
    pattern_summary = (
        result.groupby(["misclassification_pattern", "review_flag"], dropna=False)
        .size()
        .reset_index(name="row_count")
        .sort_values("row_count", ascending=False)
    )

    fixable = result[result["review_flag"].eq(1)].copy()
    human_review = result[result["review_flag"].eq(2)].copy()
    rule_agreements = result[result["cota_lot_matches_rule"] == True].copy()
    rule_disagreements = result[result["cota_lot_matches_rule"] == False].copy()

    # Legend
    legend_meaning = {
        "same_exact_regimen_repeated": "Full normalized regimen string/chunk structure is identical to previous row; likely safe auto-merge.",
        "same_drug_set_different_regimen_structure": "Same unique drugs, but bracket-level regimen structure differs; not an exact repeat.",
        "regimen_phase_drop_or_deescalation": "A regimen phase/chunk was removed or reduced while unique drugs may still look the same.",
        "drug_drop_or_deescalation": "Current unique drug set is subset of previous unique drug set; likely safe auto-merge.",
        "family_drop_or_deescalation": "Current family set is subset of previous family set; possible auto-merge.",
        "re_expansion_to_prior_regimen": "Exact regimen/chunk/drug/family combination appeared earlier in same doctor LoT; likely shrink/re-expand over-split.",
        "new_drug_addition_requires_review": "Drug was added; could be clinically meaningful, so review.",
        "same_family_drug_substitution": "Family stayed same but drug changed; review.",
        "complex_transplant_or_cell_therapy": "Transplant/CAR-T/conditioning-related case; review.",
        "complex_multi_drug_change_requires_review": "Both additions and removals; review.",
        "unknown_pattern_requires_review": "No safe rule detected; review.",
        "rule_engine_agrees_with_cota": "Rule engine and COTA agree on LOT assignment.",
        "rule_engine_disagrees_needs_review": "Rule engine disagrees with COTA; needs investigation.",
    }

    legend = pd.DataFrame([
        {
            "misclassification_pattern": pattern,
            "review_flag": 0 if "agrees" in pattern else (1 if pattern in FIXABLE_PATTERNS else 2),
            "color_hex": "#" + color,
            "meaning": legend_meaning.get(pattern, ""),
        }
        for pattern, color in PATTERN_COLORS.items()
    ])

    # Write output
    print(f"Writing output to {OUTPUT_PATH}...")
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Misclassified_Patterns", index=False)
        pattern_summary.to_excel(writer, sheet_name="Pattern_Summary", index=False)
        fixable.to_excel(writer, sheet_name="Flag_1_Fixable", index=False)
        human_review.to_excel(writer, sheet_name="Flag_2_Human_Review", index=False)
        rule_agreements.to_excel(writer, sheet_name="Rule_Agreements", index=False)
        rule_disagreements.to_excel(writer, sheet_name="Rule_Disagreements", index=False)
        legend.to_excel(writer, sheet_name="Legend", index=False)

    # Apply formatting
    print("Applying formatting...")
    wb = load_workbook(OUTPUT_PATH)
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            header = str(ws.cell(row=1, column=col_idx).value or "")
            max_len = len(header)
            for cell in column_cells[1:80]:
                max_len = max(max_len, len(str(cell.value or "")))
            width = min(max(max_len + 2, 10), 45)
            if header in {"pattern_explanation", "notes", "line_of_therapy_name"}:
                width = 42
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Color the patterns
    ws = wb["Misclassified_Patterns"]
    headers = [cell.value for cell in ws[1]]
    pattern_col_idx = headers.index("misclassification_pattern") + 1

    for row_idx in range(2, ws.max_row + 1):
        pattern = ws.cell(row=row_idx, column=pattern_col_idx).value
        fill_color = PATTERN_COLORS.get(pattern, PATTERN_COLORS["unknown_pattern_requires_review"])
        fill = PatternFill("solid", fgColor=fill_color)
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    ws_legend = wb["Legend"]
    legend_headers = [cell.value for cell in ws_legend[1]]
    legend_pattern_col = legend_headers.index("misclassification_pattern") + 1
    for row_idx in range(2, ws_legend.max_row + 1):
        pattern = ws_legend.cell(row=row_idx, column=legend_pattern_col).value
        fill_color = PATTERN_COLORS.get(pattern, PATTERN_COLORS["unknown_pattern_requires_review"])
        fill = PatternFill("solid", fgColor=fill_color)
        for col_idx in range(1, ws_legend.max_column + 1):
            ws_legend.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(OUTPUT_PATH)

    # Print summary
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total misclassified rows analyzed: {len(result)}")
    print(f"Rule engine agrees with COTA: {len(rule_agreements)} rows")
    print(f"Rule engine disagrees with COTA: {len(rule_disagreements)} rows")
    print(f"Flag 1 (likely fixable): {len(fixable)} rows")
    print(f"Flag 2 (human review): {len(human_review)} rows")
    print("\nPattern counts:")
    print(pattern_summary.to_string(index=False))
    print("\n" + "=" * 80)
    print(f"Output saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()